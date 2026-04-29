"""
Backend Flask para Inventario Ciego - Render Deploy
Conecta a Azure PostgreSQL
"""
from flask import Flask, request, jsonify, send_from_directory, send_file, render_template_string
from flask_cors import CORS
import psycopg2
from psycopg2.pool import SimpleConnectionPool
from psycopg2.extras import RealDictCursor
import os, secrets, smtplib
from decimal import Decimal
from datetime import datetime, timedelta
from io import BytesIO
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from flask.json.provider import DefaultJSONProvider

class CustomJSONProvider(DefaultJSONProvider):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)

app = Flask(__name__, static_folder='static')
app.json_provider_class = CustomJSONProvider
app.json = CustomJSONProvider(app)
CORS(app, origins=['https://inventario-ciego-5bdr.onrender.com'])

@app.after_request
def add_no_cache_headers(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# Configuracion de la base de datos Azure PostgreSQL
DB_CONFIG = {
    'host': os.environ.get('DB_HOST', 'chiosburguer.postgres.database.azure.com'),
    'database': os.environ.get('DB_NAME', 'InventariosLocales'),
    'user': os.environ.get('DB_USER', 'adminChios'),
    'password': os.environ.get('DB_PASSWORD', 'Burger2023'),
    'port': os.environ.get('DB_PORT', '5432'),
    'sslmode': 'require',
    'keepalives': 1,
    'keepalives_idle': 30,
    'keepalives_interval': 10,
    'keepalives_count': 5,
    'connect_timeout': 10
}

_connection_pool = None

def _get_pool():
    global _connection_pool
    if _connection_pool is None:
        _connection_pool = SimpleConnectionPool(
            minconn=2, maxconn=15,
            **DB_CONFIG, cursor_factory=RealDictCursor
        )
    return _connection_pool

def get_db():
    """Obtiene conexion del pool, validando que este viva"""
    conn = _get_pool().getconn()
    try:
        conn.cursor().execute("SELECT 1")
        conn.rollback()
    except Exception:
        # Conexion stale - cerrar y crear nueva
        try:
            _get_pool().putconn(conn, close=True)
        except Exception:
            try:
                conn.close()
            except Exception:
                pass
        conn = psycopg2.connect(**DB_CONFIG, cursor_factory=RealDictCursor)
    return conn

def release_db(conn):
    try:
        if conn.closed:
            return
        _get_pool().putconn(conn)
    except Exception:
        try:
            conn.close()
        except Exception:
            pass


def init_db():
    """Crea tabla merma_operativa y migra asignacion_diferencias al startup"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.merma_operativa (
                id SERIAL PRIMARY KEY,
                fecha DATE NOT NULL,
                local VARCHAR(50) NOT NULL,
                codigo VARCHAR(50) NOT NULL,
                nombre VARCHAR(150) NOT NULL,
                unidad VARCHAR(20) NOT NULL,
                cantidad NUMERIC(12,4) NOT NULL,
                motivo TEXT,
                costo_unitario NUMERIC(12,4) DEFAULT 0,
                costo_total NUMERIC(12,4) DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            ALTER TABLE goti.asignacion_diferencias
                ADD COLUMN IF NOT EXISTS codigo VARCHAR(50),
                ADD COLUMN IF NOT EXISTS nombre VARCHAR(150),
                ADD COLUMN IF NOT EXISTS unidad VARCHAR(20),
                ADD COLUMN IF NOT EXISTS local VARCHAR(50),
                ADD COLUMN IF NOT EXISTS fecha DATE
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.bajas_directas (
                id SERIAL PRIMARY KEY,
                baja_grupo BIGINT,
                fecha DATE NOT NULL,
                local VARCHAR(50) NOT NULL,
                codigo VARCHAR(50) NOT NULL,
                nombre VARCHAR(150) NOT NULL,
                unidad VARCHAR(20) NOT NULL,
                cantidad NUMERIC(12,4) NOT NULL,
                persona VARCHAR(100),
                motivo TEXT,
                costo_unitario NUMERIC(12,4) DEFAULT 0,
                costo_total NUMERIC(12,4) DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            ALTER TABLE goti.bajas_directas
                ADD COLUMN IF NOT EXISTS baja_grupo BIGINT
        """)
        cur.execute("""
            ALTER TABLE goti.bajas_directas
                ADD COLUMN IF NOT EXISTS documento VARCHAR(100)
        """)
        cur.execute("""
            ALTER TABLE goti.bajas_directas
                ADD COLUMN IF NOT EXISTS codigo_baja VARCHAR(50)
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.bajas_asignaciones (
                id SERIAL PRIMARY KEY,
                baja_grupo BIGINT NOT NULL,
                persona VARCHAR(100) NOT NULL,
                monto NUMERIC(12,2) NOT NULL,
                fecha DATE,
                local VARCHAR(50),
                motivo TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # ---- Tablas para Asignación por Sección (prototipo) ----
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.asignacion_seccion (
                id SERIAL PRIMARY KEY,
                fecha DATE NOT NULL,
                local VARCHAR(50) NOT NULL,
                nombre VARCHAR(100),
                total_valor NUMERIC(12,2) DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.asig_seccion_productos (
                id SERIAL PRIMARY KEY,
                seccion_id INT NOT NULL,
                conteo_id INT NOT NULL,
                codigo VARCHAR(50),
                nombre VARCHAR(150),
                diferencia NUMERIC(12,4),
                costo_unitario NUMERIC(12,4),
                cantidad_asignada NUMERIC(12,4),
                valor NUMERIC(12,2)
            )
        """)
        cur.execute("""
            ALTER TABLE goti.asig_seccion_productos
                ADD COLUMN IF NOT EXISTS cantidad_asignada NUMERIC(12,4)
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.asig_seccion_personas (
                id SERIAL PRIMARY KEY,
                seccion_id INT NOT NULL,
                persona VARCHAR(100),
                monto NUMERIC(12,2)
            )
        """)
        # ---- Tablas para Asignacion Semanal ----
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.semanas_inventario (
                id SERIAL PRIMARY KEY,
                fecha_inicio DATE NOT NULL,
                fecha_fin DATE NOT NULL,
                local VARCHAR(50) NOT NULL,
                estado VARCHAR(20) DEFAULT 'abierta' CHECK (estado IN ('abierta', 'cerrada')),
                cerrada_por VARCHAR(100),
                cerrada_at TIMESTAMP,
                notas TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(fecha_inicio, local)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.asignacion_semanal (
                id SERIAL PRIMARY KEY,
                semana_id INT NOT NULL,
                codigo VARCHAR(50) NOT NULL,
                nombre VARCHAR(150),
                unidad VARCHAR(20),
                local VARCHAR(50),
                diferencia_semanal NUMERIC(12,4) DEFAULT 0,
                costo_unitario NUMERIC(12,4) DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.asignacion_semanal_personas (
                id SERIAL PRIMARY KEY,
                asignacion_semanal_id INT NOT NULL,
                persona VARCHAR(100) NOT NULL,
                cantidad NUMERIC(12,4) DEFAULT 0,
                monto NUMERIC(12,2) DEFAULT 0
            )
        """)
        # ---- Columnas de auditoria: quien contó y quien modificó ----
        cur.execute("""
            ALTER TABLE goti.inventario_ciego_conteos
                ADD COLUMN IF NOT EXISTS contado_por VARCHAR(50),
                ADD COLUMN IF NOT EXISTS contado_at TIMESTAMP,
                ADD COLUMN IF NOT EXISTS contado2_por VARCHAR(50),
                ADD COLUMN IF NOT EXISTS contado2_at TIMESTAMP,
                ADD COLUMN IF NOT EXISTS modificado_por VARCHAR(50),
                ADD COLUMN IF NOT EXISTS modificado_at TIMESTAMP
        """)
        # ---- Tabla de permisos por ROL (ver + editar) ----
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.rol_modulos (
                id SERIAL PRIMARY KEY,
                rol VARCHAR(20) NOT NULL,
                modulo VARCHAR(30) NOT NULL,
                puede_ver BOOLEAN DEFAULT TRUE,
                puede_editar BOOLEAN DEFAULT FALSE,
                UNIQUE(rol, modulo)
            )
        """)
        # Migrar: agregar columnas si tabla ya existia sin ellas
        cur.execute("""
            ALTER TABLE goti.rol_modulos
                ADD COLUMN IF NOT EXISTS puede_ver BOOLEAN DEFAULT TRUE,
                ADD COLUMN IF NOT EXISTS puede_editar BOOLEAN DEFAULT FALSE,
                ADD COLUMN IF NOT EXISTS puede_eliminar BOOLEAN DEFAULT FALSE
        """)
        # Seed defaults si la tabla esta vacia
        cur.execute("SELECT COUNT(*) as cnt FROM goti.rol_modulos")
        if cur.fetchone()['cnt'] == 0:
            # Subgerente: conteo, observaciones, historico, dashboard
            for mod in ['conteo','observaciones','historico','dashboard']:
                cur.execute("INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar) VALUES ('subgerente', %s, TRUE, TRUE, FALSE) ON CONFLICT DO NOTHING", (mod,))
            # Supervisor: ve todos los locales, ve todo pero no edita usuarios
            for mod in ['conteo','observaciones','historico','dashboard','cruce','bajas','semanal','correccion']:
                cur.execute("INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar) VALUES ('supervisor', %s, TRUE, TRUE, FALSE) ON CONFLICT DO NOTHING", (mod,))
            # Gerente: todo lo del subgerente + semanal, cruce, bajas
            for mod in ['conteo','observaciones','historico','dashboard']:
                cur.execute("INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar) VALUES ('gerente', %s, TRUE, TRUE, FALSE) ON CONFLICT DO NOTHING", (mod,))
            for mod in ['cruce','bajas','semanal','correccion']:
                cur.execute("INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar) VALUES ('gerente', %s, TRUE, TRUE, FALSE) ON CONFLICT DO NOTHING", (mod,))
            # Admin: ve, edita y elimina todo
            for mod in ['conteo','observaciones','historico','dashboard','cruce','bajas','semanal','correccion','usuarios']:
                cur.execute("INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar) VALUES ('admin', %s, TRUE, TRUE, TRUE) ON CONFLICT DO NOTHING", (mod,))
        # Actualizar registros existentes que no tengan puede_ver seteado (migracion)
        cur.execute("UPDATE goti.rol_modulos SET puede_ver = TRUE WHERE puede_ver IS NULL")
        cur.execute("UPDATE goti.rol_modulos SET puede_editar = TRUE WHERE puede_editar IS NULL")

        # Migrar roles: empleado → subgerente, supervisor → gerente
        cur.execute("UPDATE goti.usuarios SET rol = 'subgerente' WHERE rol = 'empleado'")
        cur.execute("UPDATE goti.usuarios SET rol = 'gerente' WHERE rol = 'supervisor'")
        cur.execute("UPDATE goti.rol_modulos SET rol = 'subgerente' WHERE rol = 'empleado'")
        cur.execute("UPDATE goti.rol_modulos SET rol = 'gerente' WHERE rol = 'supervisor'")
        conn.commit()
        print('init_db: tablas OK')
    except Exception as e:
        print(f'init_db error: {e}')
    finally:
        if conn:
            release_db(conn)


# Helper: mapeo de IDs de bodega a nombres legibles
BODEGAS_NOMBRES = {
    'real_audiencia': 'Real Audiencia',
    'floreana': 'Floreana',
    'portugal': 'Portugal',
    'santo_cachon_real': 'Santo Cachon Real',
    'santo_cachon_portugal': 'Santo Cachon Portugal',
    'simon_bolon': 'Simon Bolon',
    'bodega_principal': 'Bodega Principal',
    'materia_prima': 'Materia Prima',
    'planta': 'Planta de Produccion'
}

# (USUARIO_BODEGA eliminado — permisos de bodega ahora se manejan desde BD tabla usuario_bodegas)

# ==================== RUTAS ESTATICAS ====================

@app.route('/')
def index():
    import json as json_lib, base64
    # Inyectar personas directamente en el HTML como JSON en data attribute (evita problemas de encoding en script)
    personas = _personas_cache['datos'] if _personas_cache['datos'] else []
    if not personas:
        try:
            personas = _cargar_personas_airtable()
        except Exception:
            pass
    html_path = os.path.join(app.static_folder, 'index.html')
    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()
    # Usar base64 para evitar cualquier problema de encoding/caracteres especiales
    personas_json = json_lib.dumps(personas, ensure_ascii=True)
    personas_b64 = base64.b64encode(personas_json.encode('utf-8')).decode('ascii')
    inject = f'<script id="personas-data" type="application/json">{personas_json}</script>\n'
    inject += f'<meta name="personas-b64" content="{personas_b64}">\n'
    html = html.replace('</head>', inject + '</head>')
    return html

@app.route('/establecer-clave')
def pagina_establecer_clave():
    """Pagina publica donde el usuario establece su contrasena."""
    token = request.args.get('token', '')
    if not token:
        return PAGINA_TOKEN_INVALIDO, 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""SELECT username, nombre, invite_token_expires FROM goti.usuarios
                       WHERE invite_token = %s AND activo = TRUE""", (token,))
        user = cur.fetchone()
        if not user:
            return PAGINA_TOKEN_INVALIDO, 404
        if user['invite_token_expires'] and user['invite_token_expires'] < datetime.utcnow():
            return PAGINA_TOKEN_INVALIDO, 410
        html = PAGINA_ESTABLECER_CLAVE.replace('{{ nombre }}', user['nombre']).replace('{{ username }}', user['username']).replace('{{ token }}', token)
        return html
    except Exception as e:
        print(f"Error en /establecer-clave: {e}")
        return PAGINA_TOKEN_INVALIDO, 500
    finally:
        if conn:
            release_db(conn)


@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('static', path)

# ==================== API ====================

_login_attempts = {}

def _check_rate_limit(ip, max_attempts=5, window=60):
    now = _time.time()
    attempts = _login_attempts.get(ip, [])
    attempts = [t for t in attempts if now - t < window]
    _login_attempts[ip] = attempts
    return len(attempts) < max_attempts

def _record_login_attempt(ip):
    now = _time.time()
    if ip not in _login_attempts:
        _login_attempts[ip] = []
    _login_attempts[ip].append(now)

@app.route('/api/login', methods=['POST'])
def login():
    ip = request.remote_addr
    if not _check_rate_limit(ip):
        return jsonify({'success': False, 'error': 'Demasiados intentos. Espera 60 segundos.'}), 429

    data = request.json
    username = data.get('username')
    password = data.get('password')

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT username, nombre, rol FROM goti.usuarios
            WHERE username = %s AND password = %s AND activo = TRUE
        """, (username, password))
        user = cur.fetchone()

        if user:
            # Cargar bodegas desde BD
            cur.execute("""
                SELECT ub.bodega FROM goti.usuario_bodegas ub
                JOIN goti.usuarios u ON u.id = ub.usuario_id
                WHERE u.username = %s
                ORDER BY ub.bodega
            """, (user['username'],))
            bodegas_user = [r['bodega'] for r in cur.fetchall()]
            # Compatibilidad: si tiene 1 sola bodega de ventas, enviar como string
            bodegas_ventas = [b for b in bodegas_user if b not in ('bodega_principal', 'materia_prima', 'planta')]
            bodega_asignada = bodegas_ventas[0] if len(bodegas_ventas) == 1 else None
            # Cargar modulos permitidos segun el ROL (con nivel ver/editar)
            cur.execute("""
                SELECT modulo, puede_ver, puede_editar FROM goti.rol_modulos
                WHERE rol = %s ORDER BY modulo
            """, (user['rol'],))
            modulos_user = [r['modulo'] for r in cur.fetchall() if r['puede_ver']]
            permisos_user = {}
            cur.execute("""
                SELECT modulo, puede_ver, puede_editar, COALESCE(puede_eliminar, FALSE) as puede_eliminar
                FROM goti.rol_modulos WHERE rol = %s
            """, (user['rol'],))
            for r in cur.fetchall():
                permisos_user[r['modulo']] = {'ver': r['puede_ver'], 'editar': r['puede_editar'], 'eliminar': r['puede_eliminar']}
            return jsonify({
                'success': True,
                'user': {
                    'username': user['username'],
                    'nombre': user['nombre'],
                    'rol': user['rol'],
                    'bodega': bodega_asignada,
                    'bodegas': bodegas_user,
                    'modulos': modulos_user,
                    'permisos': permisos_user
                }
            })

        _record_login_attempt(ip)
        return jsonify({'success': False, 'error': 'Credenciales invalidas'}), 401
    except Exception as e:
        print(f"Error en /api/login: {e}")
        return jsonify({'success': False, 'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/categorias', methods=['GET'])
def get_categorias():
    # Categorias estaticas
    categorias = [
        {'id': 1, 'nombre': 'Bebidas'},
        {'id': 2, 'nombre': 'Carnes'},
        {'id': 3, 'nombre': 'Lacteos'},
        {'id': 4, 'nombre': 'Congelados'},
        {'id': 5, 'nombre': 'Otros'}
    ]
    return jsonify(categorias)

@app.route('/api/bodegas', methods=['GET'])
def get_bodegas():
    bodegas = [
        {'id': 'real_audiencia', 'nombre': 'Real Audiencia'},
        {'id': 'floreana', 'nombre': 'Floreana'},
        {'id': 'portugal', 'nombre': 'Portugal'},
        {'id': 'santo_cachon_real', 'nombre': 'Santo Cachon Real'},
        {'id': 'santo_cachon_portugal', 'nombre': 'Santo Cachon Portugal'},
        {'id': 'simon_bolon', 'nombre': 'Simon Bolon'},
        {'id': 'bodega_principal', 'nombre': 'Bodega Principal'},
        {'id': 'materia_prima', 'nombre': 'Materia Prima'},
        {'id': 'planta', 'nombre': 'Planta de Produccion'}
    ]
    return jsonify(bodegas)

@app.route('/api/inventario/consultar', methods=['GET'])
def consultar_inventario():
    fecha = request.args.get('fecha')
    local = request.args.get('local')

    if not fecha or not local:
        return jsonify({'error': 'Fecha y local son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Asegurar columnas: observaciones, motivo, corregido (auditoría) y justificado (no descontar)
        cur.execute("""
            ALTER TABLE goti.inventario_ciego_conteos
            ADD COLUMN IF NOT EXISTS observaciones TEXT;
            ALTER TABLE goti.inventario_ciego_conteos
            ADD COLUMN IF NOT EXISTS motivo TEXT;
            ALTER TABLE goti.inventario_ciego_conteos
            ADD COLUMN IF NOT EXISTS corregido BOOLEAN DEFAULT FALSE;
            ALTER TABLE goti.inventario_ciego_conteos
            ADD COLUMN IF NOT EXISTS justificado BOOLEAN DEFAULT FALSE;
        """)
        conn.commit()

        cur.execute("""
            SELECT id, codigo, nombre, unidad, cantidad, cantidad_contada, cantidad_contada_2, observaciones,
                   COALESCE(motivo, '') as motivo,
                   COALESCE(corregido, FALSE) as corregido,
                   COALESCE(justificado, FALSE) as justificado,
                   COALESCE(costo_unitario, 0) as costo_unitario
            FROM goti.inventario_ciego_conteos
            WHERE fecha = %s AND local = %s
            ORDER BY codigo
        """, (fecha, local))

        productos = cur.fetchall()

        # Incluir personas del cache (nunca bloquea, solo datos en memoria)
        personas = _personas_cache['datos']

        return jsonify({'productos': productos, 'personas': personas})
    except Exception as e:
        print(f"Error en /api/inventario/consultar: {e}")
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/inventario/autofill-conteo2', methods=['POST'])
def autofill_conteo2():
    """Auto-llena conteo 2 con conteo 1 para productos donde conteo1 == sistema"""
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')

    if not fecha or not local:
        return jsonify({'error': 'fecha y local son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.inventario_ciego_conteos
            SET cantidad_contada_2 = cantidad_contada
            WHERE fecha = %s AND local = %s
              AND cantidad_contada IS NOT NULL
              AND cantidad_contada_2 IS NULL
              AND cantidad_contada = cantidad
        """, (fecha, local))
        actualizados = cur.rowcount
        conn.commit()

        return jsonify({'success': True, 'actualizados': actualizados})
    except Exception as e:
        print(f"Error en /api/inventario/autofill-conteo2: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/inventario/guardar-conteo', methods=['POST'])
def guardar_conteo():
    data = request.json
    id_producto = data.get('id')
    cantidad = data.get('cantidad_contada')
    conteo = data.get('conteo', 1)
    usuario = data.get('usuario', '')

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        if conteo == 2:
            cur.execute("""
                UPDATE goti.inventario_ciego_conteos
                SET cantidad_contada_2 = %s, contado2_por = %s, contado2_at = NOW()
                WHERE id = %s
            """, (cantidad, usuario or None, id_producto))
        else:
            cur.execute("""
                UPDATE goti.inventario_ciego_conteos
                SET cantidad_contada = %s, contado_por = %s, contado_at = NOW()
                WHERE id = %s
            """, (cantidad, usuario or None, id_producto))

        conn.commit()

        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en /api/inventario/guardar-conteo: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/inventario/guardar-observacion', methods=['POST'])
def guardar_observacion():
    data = request.json
    id_producto = data.get('id')
    observaciones = data.get('observaciones', None)
    motivo = data.get('motivo', None)
    corregido = data.get('corregido', None)
    justificado = data.get('justificado', None)

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Construir SET dinámico según los campos enviados
        sets = []
        params = []
        if observaciones is not None:
            sets.append("observaciones = %s")
            params.append(observaciones)
        if motivo is not None:
            sets.append("motivo = %s")
            params.append(motivo)
        if corregido is not None:
            sets.append("corregido = %s")
            params.append(bool(corregido))
        if justificado is not None:
            sets.append("justificado = %s")
            params.append(bool(justificado))

        if sets:
            params.append(id_producto)
            cur.execute(f"""
                UPDATE goti.inventario_ciego_conteos
                SET {', '.join(sets)}
                WHERE id = %s
            """, params)
        conn.commit()

        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en /api/inventario/guardar-observacion: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)

# ==================== REPORTE MOTIVOS ====================

@app.route('/api/reportes/motivos-lista', methods=['GET'])
def reporte_motivos_lista():
    """Devuelve lista de motivos unicos disponibles."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT motivo FROM goti.inventario_ciego_conteos
            WHERE motivo IS NOT NULL AND motivo != ''
            ORDER BY motivo
        """)
        motivos = [r['motivo'] for r in cur.fetchall()]
        return jsonify(motivos)
    except Exception as e:
        return jsonify([])
    finally:
        if conn: release_db(conn)


@app.route('/api/reportes/motivos', methods=['GET'])
def reporte_motivos():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodegas = request.args.getlist('bodega')
    bodegas = [b for b in bodegas if b]
    producto = request.args.get('producto', '')

    if not fecha_desde or not fecha_hasta:
        return jsonify({'error': 'fecha_desde y fecha_hasta requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Asegurar columna motivo existe en conteos
        cur.execute("""
            ALTER TABLE goti.inventario_ciego_conteos
            ADD COLUMN IF NOT EXISTS motivo TEXT
        """)
        conn.commit()

        # Asegurar tabla manuales existe
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.observaciones_manuales (
                id SERIAL PRIMARY KEY,
                fecha DATE NOT NULL,
                local VARCHAR(100) NOT NULL,
                codigo VARCHAR(50),
                nombre VARCHAR(255) NOT NULL,
                diferencia NUMERIC(12,3) DEFAULT 0,
                motivo TEXT,
                observaciones TEXT,
                corregido BOOLEAN DEFAULT FALSE,
                creado_por VARCHAR(100),
                creado_at TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit()

        # Motivos de conteos
        query1 = """
            SELECT motivo, COUNT(*) as cantidad
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s
              AND motivo IS NOT NULL AND motivo != ''
        """
        params1 = [fecha_desde, fecha_hasta]
        if producto:
            query1 += " AND codigo = %s"
            params1.append(producto)
        if len(bodegas) == 1:
            query1 += " AND local = %s"
            params1.append(bodegas[0])
        elif len(bodegas) > 1:
            query1 += " AND local IN (" + ",".join(["%s"] * len(bodegas)) + ")"
            params1.extend(bodegas)
        query1 += " GROUP BY motivo"

        cur.execute(query1, params1)
        motivos_conteo = cur.fetchall()

        # Motivos de observaciones manuales
        query2 = """
            SELECT motivo, COUNT(*) as cantidad
            FROM goti.observaciones_manuales
            WHERE fecha >= %s AND fecha <= %s
              AND motivo IS NOT NULL AND motivo != ''
        """
        params2 = [fecha_desde, fecha_hasta]
        if producto:
            query2 += " AND codigo = %s"
            params2.append(producto)
        if len(bodegas) == 1:
            query2 += " AND local = %s"
            params2.append(bodegas[0])
        elif len(bodegas) > 1:
            query2 += " AND local IN (" + ",".join(["%s"] * len(bodegas)) + ")"
            params2.extend(bodegas)
        query2 += " GROUP BY motivo"

        cur.execute(query2, params2)
        motivos_manual = cur.fetchall()

        # Combinar ambos
        totales = {}
        for m in motivos_conteo:
            totales[m['motivo']] = totales.get(m['motivo'], 0) + m['cantidad']
        for m in motivos_manual:
            totales[m['motivo']] = totales.get(m['motivo'], 0) + m['cantidad']

        resultado = [{'motivo': k, 'cantidad': v} for k, v in totales.items()]
        resultado.sort(key=lambda x: x['cantidad'], reverse=True)

        return jsonify(resultado)
    except Exception as e:
        print(f"Error en /api/reportes/motivos: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/reportes/diferencias-fecha', methods=['GET'])
def reporte_diferencias_fecha():
    """Productos con diferencia para una fecha y bodega específica"""
    fecha = request.args.get('fecha')
    bodega = request.args.get('bodega', '')

    if not fecha:
        return jsonify({'error': 'fecha requerida'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT nombre, unidad,
                   cantidad as sistema,
                   COALESCE(cantidad_contada_2, cantidad_contada) as conteo,
                   COALESCE(cantidad_contada_2, cantidad_contada) - cantidad as diferencia,
                   COALESCE(motivo, '') as motivo
            FROM goti.inventario_ciego_conteos
            WHERE fecha = %s
              AND COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
              AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
        """
        params = [fecha]
        if bodega:
            query += " AND local = %s"
            params.append(bodega)
        query += " ORDER BY ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad) DESC LIMIT 20"

        cur.execute(query, params)
        productos = [{
            'nombre': r['nombre'],
            'unidad': r['unidad'],
            'sistema': float(r['sistema']),
            'conteo': float(r['conteo']),
            'diferencia': float(r['diferencia']),
            'motivo': r['motivo']
        } for r in cur.fetchall()]

        return jsonify(productos)
    except Exception as e:
        print(f"Error en /api/reportes/diferencias-fecha: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/reportes/motivos/detalle', methods=['GET'])
def reporte_motivo_detalle():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    motivo = request.args.get('motivo', '')
    bodega = request.args.get('bodega', '')

    if not fecha_desde or not fecha_hasta or not motivo:
        return jsonify({'error': 'fecha_desde, fecha_hasta y motivo requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Productos de conteos con ese motivo
        query1 = """
            SELECT nombre, local, COUNT(*) as veces,
                   SUM(ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad)) as diferencia_total
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s AND motivo = %s
        """
        params1 = [fecha_desde, fecha_hasta, motivo]
        if bodega:
            query1 += " AND local = %s"
            params1.append(bodega)
        query1 += " GROUP BY nombre, local ORDER BY veces DESC, diferencia_total DESC"
        cur.execute(query1, params1)
        productos_conteo = cur.fetchall()

        # Productos de observaciones manuales con ese motivo
        query2 = """
            SELECT nombre, local, COUNT(*) as veces,
                   SUM(ABS(diferencia)) as diferencia_total
            FROM goti.observaciones_manuales
            WHERE fecha >= %s AND fecha <= %s AND motivo = %s
        """
        params2 = [fecha_desde, fecha_hasta, motivo]
        if bodega:
            query2 += " AND local = %s"
            params2.append(bodega)
        query2 += " GROUP BY nombre, local ORDER BY veces DESC"
        cur.execute(query2, params2)
        productos_manual = cur.fetchall()

        # Combinar
        combinado = {}
        for p in productos_conteo:
            key = p['nombre']
            if key not in combinado:
                combinado[key] = {'nombre': p['nombre'], 'veces': 0, 'diferencia_total': 0, 'bodegas': set()}
            combinado[key]['veces'] += p['veces']
            combinado[key]['diferencia_total'] += float(p['diferencia_total'] or 0)
            combinado[key]['bodegas'].add(BODEGAS_NOMBRES.get(p['local'], p['local']))

        for p in productos_manual:
            key = p['nombre']
            if key not in combinado:
                combinado[key] = {'nombre': p['nombre'], 'veces': 0, 'diferencia_total': 0, 'bodegas': set()}
            combinado[key]['veces'] += p['veces']
            combinado[key]['diferencia_total'] += float(p['diferencia_total'] or 0)
            combinado[key]['bodegas'].add(BODEGAS_NOMBRES.get(p['local'], p['local']))

        resultado = []
        for v in combinado.values():
            resultado.append({
                'nombre': v['nombre'],
                'veces': v['veces'],
                'diferencia_total': round(v['diferencia_total'], 3),
                'bodegas': ', '.join(sorted(v['bodegas']))
            })
        resultado.sort(key=lambda x: x['veces'], reverse=True)

        return jsonify(resultado)
    except Exception as e:
        print(f"Error en /api/reportes/motivos/detalle: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

# ==================== OBSERVACIONES MANUALES ====================

@app.route('/api/observaciones-manuales', methods=['GET'])
def listar_obs_manuales():
    fecha = request.args.get('fecha')
    local = request.args.get('local')
    if not fecha or not local:
        return jsonify({'error': 'fecha y local requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.observaciones_manuales (
                id SERIAL PRIMARY KEY,
                fecha DATE NOT NULL,
                local VARCHAR(100) NOT NULL,
                codigo VARCHAR(50),
                nombre VARCHAR(255) NOT NULL,
                diferencia NUMERIC(12,3) DEFAULT 0,
                motivo TEXT,
                observaciones TEXT,
                corregido BOOLEAN DEFAULT FALSE,
                justificado BOOLEAN DEFAULT FALSE,
                creado_por VARCHAR(100),
                creado_at TIMESTAMP DEFAULT NOW()
            )
        """)
        # Asegurar columna justificado (migracion)
        cur.execute("ALTER TABLE goti.observaciones_manuales ADD COLUMN IF NOT EXISTS justificado BOOLEAN DEFAULT FALSE")
        conn.commit()

        cur.execute("""
            SELECT id, codigo, nombre, diferencia, motivo, observaciones, corregido, COALESCE(justificado, FALSE) as justificado, creado_por
            FROM goti.observaciones_manuales
            WHERE fecha = %s AND local = %s
            ORDER BY creado_at
        """, (fecha, local))
        return jsonify(cur.fetchall())
    except Exception as e:
        print(f"Error en /api/observaciones-manuales GET: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/observaciones-manuales', methods=['POST'])
def crear_obs_manual():
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')
    codigo = data.get('codigo', '')
    nombre = data.get('nombre', '')
    diferencia = data.get('diferencia', 0)
    motivo = data.get('motivo', '')
    observaciones = data.get('observaciones', '')
    creado_por = data.get('creado_por', '')

    if not fecha or not local or not nombre:
        return jsonify({'error': 'fecha, local y nombre son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.observaciones_manuales
            (fecha, local, codigo, nombre, diferencia, motivo, observaciones, creado_por)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (fecha, local, codigo, nombre, float(diferencia), motivo, observaciones, creado_por))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        print(f"Error en /api/observaciones-manuales POST: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/observaciones-manuales/<int:obs_id>', methods=['PUT'])
def actualizar_obs_manual(obs_id):
    data = request.json
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        sets = []
        params = []
        for campo in ['motivo', 'observaciones', 'diferencia']:
            if campo in data:
                sets.append(f"{campo} = %s")
                params.append(data[campo])
        if 'corregido' in data:
            sets.append("corregido = %s")
            params.append(bool(data['corregido']))
        if 'justificado' in data:
            sets.append("justificado = %s")
            params.append(bool(data['justificado']))
        if sets:
            params.append(obs_id)
            cur.execute(f"""
                UPDATE goti.observaciones_manuales
                SET {', '.join(sets)}
                WHERE id = %s
            """, params)
            conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en /api/observaciones-manuales PUT: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/observaciones-manuales/<int:obs_id>', methods=['DELETE'])
def eliminar_obs_manual(obs_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.observaciones_manuales WHERE id = %s", (obs_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en /api/observaciones-manuales DELETE: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/admin/corregir-conteo', methods=['PUT'])
def corregir_conteo():
    """Permite al admin corregir conteo1 y/o conteo2 de un producto"""
    data = request.json
    id_producto = data.get('id')
    cantidad_contada = data.get('cantidad_contada')
    cantidad_contada_2 = data.get('cantidad_contada_2')
    cantidad_sistema = data.get('cantidad')
    usuario = data.get('usuario', '')

    if id_producto is None:
        return jsonify({'error': 'id es requerido'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.inventario_ciego_conteos
            SET cantidad = COALESCE(%s, cantidad),
                cantidad_contada = %s,
                cantidad_contada_2 = %s,
                modificado_por = %s,
                modificado_at = CURRENT_TIMESTAMP,
                corregido = TRUE
            WHERE id = %s
        """, (cantidad_sistema, cantidad_contada, cantidad_contada_2, usuario or None, id_producto))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en /api/admin/corregir-conteo: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/agregar-producto', methods=['POST'])
def admin_agregar_producto():
    """Agrega un producto a una fecha+bodega con cantidad 0"""
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')
    codigo = data.get('codigo', '').strip().upper()
    nombre = data.get('nombre', '').strip()
    unidad = data.get('unidad', 'Unidad').strip()

    if not fecha or not local or not codigo or not nombre:
        return jsonify({'error': 'fecha, local, codigo y nombre son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Verificar que no exista ya
        cur.execute("""
            SELECT id FROM goti.inventario_ciego_conteos
            WHERE fecha = %s AND local = %s AND codigo = %s
        """, (fecha, local, codigo))
        if cur.fetchone():
            return jsonify({'error': f'El producto {codigo} ya existe para esta fecha y bodega'}), 409

        cur.execute("""
            INSERT INTO goti.inventario_ciego_conteos
                (fecha, local, codigo, nombre, unidad, cantidad, costo_unitario)
            VALUES (%s, %s, %s, %s, %s, 0, 0)
            RETURNING id
        """, (fecha, local, codigo, nombre, unidad))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        print(f"Error en /api/admin/agregar-producto: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/eliminar-producto', methods=['DELETE'])
def admin_eliminar_producto():
    """Elimina un producto de una fecha+bodega"""
    data = request.json
    id_producto = data.get('id')

    if not id_producto:
        return jsonify({'error': 'id es requerido'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            DELETE FROM goti.inventario_ciego_conteos WHERE id = %s
        """, (id_producto,))
        if cur.rowcount == 0:
            return jsonify({'error': 'Producto no encontrado'}), 404
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en /api/admin/eliminar-producto: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


# ============================================================
# MODULO: Configuracion de Productos por Marca
# ============================================================

@app.route('/api/admin/productos-marca', methods=['GET'])
def get_productos_marca():
    """Lista productos configurados para una marca"""
    marca = request.args.get('marca', '').upper()
    if not marca:
        return jsonify({'error': 'marca es requerido'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, codigo, nombre, activo
            FROM goti.productos_por_marca
            WHERE marca = %s
            ORDER BY codigo
        """, (marca,))
        productos = [{'id': r['id'], 'codigo': r['codigo'], 'nombre': r['nombre'], 'activo': r['activo']} for r in cur.fetchall()]
        return jsonify(productos)
    except Exception as e:
        print(f"Error en get_productos_marca: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/productos-marca', methods=['POST'])
def add_producto_marca():
    """Agrega un producto a una marca"""
    data = request.json
    marca = data.get('marca', '').upper()
    codigo = data.get('codigo', '').strip().upper()
    nombre = data.get('nombre', '').strip()
    if not marca or not codigo or not nombre:
        return jsonify({'error': 'marca, codigo y nombre son requeridos'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.productos_por_marca (marca, codigo, nombre, activo)
            VALUES (%s, %s, %s, TRUE)
            ON CONFLICT (marca, codigo) DO UPDATE SET nombre = EXCLUDED.nombre, activo = TRUE
            RETURNING id
        """, (marca, codigo, nombre))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        print(f"Error en add_producto_marca: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/productos-marca/<int:pid>', methods=['DELETE'])
def delete_producto_marca(pid):
    """Elimina un producto de una marca"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.productos_por_marca WHERE id = %s", (pid,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en delete_producto_marca: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/productos-marca/toggle/<int:pid>', methods=['PUT'])
def toggle_producto_marca(pid):
    """Activa/desactiva un producto"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.productos_por_marca SET activo = NOT activo WHERE id = %s
            RETURNING activo
        """, (pid,))
        row = cur.fetchone()
        conn.commit()
        return jsonify({'success': True, 'activo': row['activo'] if row else False})
    except Exception as e:
        print(f"Error en toggle_producto_marca: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/inventario/cargar', methods=['POST'])
def cargar_inventario():
    """Endpoint para cargar datos desde el script de Selenium"""
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')
    productos = data.get('productos', [])

    if not fecha or not local or not productos:
        return jsonify({'error': 'Datos incompletos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        registros = 0
        for prod in productos:
            cur.execute("""
                INSERT INTO goti.inventario_ciego_conteos
                (fecha, local, codigo, nombre, unidad, cantidad)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (fecha, local, codigo)
                DO UPDATE SET cantidad = EXCLUDED.cantidad, nombre = EXCLUDED.nombre
            """, (fecha, local, prod['codigo'], prod['nombre'], prod['unidad'], prod['cantidad']))
            registros += 1

        conn.commit()

        return jsonify({'success': True, 'registros': registros})
    except Exception as e:
        print(f"Error en /api/inventario/cargar: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/inventario/generar-conteo-operativo', methods=['POST'])
def generar_conteo_operativo():
    """Crea tarea para que el worker descargue stock de Contifico y genere conteo.
    - Bodega Principal: 8 fijos + 5 aleatorios semanales
    - Materia Prima / Planta: 10 aleatorios semanales
    Body: {bodega, fecha}"""
    data = request.json or {}
    bodega = data.get('bodega')
    fecha = data.get('fecha')

    BODEGAS_VALIDAS = ('bodega_principal', 'materia_prima')
    if bodega not in BODEGAS_VALIDAS:
        return jsonify({'error': 'bodega invalida'}), 400
    if not fecha:
        return jsonify({'error': 'fecha requerida'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Verificar si ya existen productos para esta fecha+bodega
        cur.execute("""
            SELECT COUNT(*) as n FROM goti.inventario_ciego_conteos
            WHERE fecha = %s AND local = %s
        """, (fecha, bodega))
        ya_existe = cur.fetchone()['n']
        if ya_existe > 0:
            return jsonify({'error': f'Ya existen {ya_existe} productos para {bodega} en {fecha}. No se puede regenerar.', 'ya_existe': True}), 409

        # Crear tarea para el worker
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.conteo_operativo_tareas (
                id SERIAL PRIMARY KEY,
                bodega VARCHAR(50) NOT NULL,
                fecha DATE NOT NULL,
                estado VARCHAR(20) DEFAULT 'pendiente',
                solicitado_at TIMESTAMP DEFAULT NOW(),
                worker_lock VARCHAR(50),
                timestamp_inicio TIMESTAMP,
                timestamp_fin TIMESTAMP,
                total_productos INT,
                fijos INT,
                aleatorios INT,
                error_msg TEXT,
                UNIQUE(bodega, fecha)
            )
        """)
        conn.commit()

        # Verificar si ya hay tarea pendiente/en_proceso
        cur.execute("""
            SELECT id, estado FROM goti.conteo_operativo_tareas
            WHERE bodega = %s AND fecha = %s
        """, (bodega, fecha))
        existente = cur.fetchone()
        if existente:
            if existente['estado'] in ('pendiente', 'en_proceso'):
                return jsonify({'id': existente['id'], 'estado': existente['estado'], 'reused': True})
            if existente['estado'] == 'completado':
                return jsonify({'error': 'Ya se genero el conteo para esta fecha', 'ya_existe': True}), 409
            # Error: resetear
            cur.execute("""
                UPDATE goti.conteo_operativo_tareas
                SET estado='pendiente', solicitado_at=NOW(), worker_lock=NULL, error_msg=NULL,
                    timestamp_inicio=NULL, timestamp_fin=NULL, total_productos=NULL
                WHERE id = %s
            """, (existente['id'],))
            conn.commit()
            return jsonify({'id': existente['id'], 'estado': 'pendiente', 'reset': True})

        cur.execute("""
            INSERT INTO goti.conteo_operativo_tareas (bodega, fecha)
            VALUES (%s, %s) RETURNING id
        """, (bodega, fecha))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'id': new_id, 'estado': 'pendiente'})
    except Exception as e:
        print(f"Error en generar-conteo-operativo: {e}")
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/conteo-op/pendientes', methods=['GET'])
def conteo_op_pendientes():
    """Worker toma tareas de conteo operativo."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401
    worker_id = request.args.get('worker_id', 'pc-finanzas')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.conteo_operativo_tareas
            SET estado = 'en_proceso', worker_lock = %s, timestamp_inicio = NOW()
            WHERE id IN (
                SELECT id FROM goti.conteo_operativo_tareas
                WHERE estado = 'pendiente' ORDER BY solicitado_at ASC LIMIT 1
                FOR UPDATE SKIP LOCKED
            )
            RETURNING id, bodega, fecha
        """, (worker_id,))
        rows = cur.fetchall()
        conn.commit()
        return jsonify([{
            'id': r['id'], 'bodega': r['bodega'],
            'fecha': r['fecha'].isoformat() if r['fecha'] else None,
            'tipo': 'conteo_operativo',
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/conteo-op/resultado', methods=['POST'])
def conteo_op_resultado():
    """Worker reporta resultado del conteo operativo."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401
    data = request.json or {}
    ejec_id = data.get('id')
    estado = data.get('estado', 'completado')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.conteo_operativo_tareas
            SET estado = %s, timestamp_fin = NOW(),
                total_productos = %s, fijos = %s, aleatorios = %s, error_msg = %s
            WHERE id = %s
        """, (estado, data.get('total_productos'), data.get('fijos'), data.get('aleatorios'),
              data.get('error_msg'), ejec_id))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/conteo-op/estado/<int:ejec_id>', methods=['GET'])
def conteo_op_estado(ejec_id):
    """Polling del panel."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM goti.conteo_operativo_tareas WHERE id = %s", (ejec_id,))
        r = cur.fetchone()
        if not r: return jsonify({'error': 'no encontrado'}), 404
        return jsonify({
            'id': r['id'], 'bodega': r['bodega'], 'estado': r['estado'],
            'total_productos': r['total_productos'], 'fijos': r['fijos'],
            'aleatorios': r['aleatorios'], 'error_msg': r['error_msg'],
        })
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/historico', methods=['GET'])
def historico():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodega = request.args.get('bodega')

    if not fecha_desde or not fecha_hasta:
        return jsonify({'error': 'fecha_desde y fecha_hasta son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT
                fecha,
                local,
                COUNT(*) as total_productos,
                COUNT(cantidad_contada) as total_contados,
                COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
                    THEN 1 END) as total_con_diferencia,
                COUNT(CASE WHEN cantidad_contada IS NOT NULL THEN 1 END) as total_con_conteo1,
                COUNT(CASE WHEN cantidad_contada_2 IS NOT NULL THEN 1 END) as total_con_conteo2
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s
        """
        params = [fecha_desde, fecha_hasta]

        if bodega:
            query += " AND local = %s"
            params.append(bodega)

        query += " GROUP BY fecha, local ORDER BY fecha DESC, local"

        cur.execute(query, params)
        resultados = cur.fetchall()

        # Calcular estado para cada registro
        datos = []
        for r in resultados:
            total = r['total_productos']
            contados = r['total_contados']
            con_conteo2 = r['total_con_conteo2']

            if con_conteo2 > 0 or (contados == total and r['total_con_diferencia'] == 0):
                estado = 'completo'
            elif contados > 0:
                estado = 'en_proceso'
            else:
                estado = 'pendiente'

            porcentaje = round((contados / total * 100) if total > 0 else 0)

            datos.append({
                'fecha': str(r['fecha']),
                'local': r['local'],
                'total_productos': total,
                'total_contados': contados,
                'total_con_diferencia': r['total_con_diferencia'],
                'estado': estado,
                'porcentaje': porcentaje
            })

        return jsonify(datos)
    except Exception as e:
        print(f"Error en /api/historico: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/historico/pivot', methods=['GET'])
def historico_pivot():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    local = request.args.get('bodega')
    if not fecha_desde or not fecha_hasta or not local:
        return jsonify({'error': 'fecha_desde, fecha_hasta y bodega son requeridos'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT
                c.id, c.codigo, c.nombre, c.unidad,
                c.fecha,
                c.cantidad AS stock,
                COALESCE(c.cantidad_contada_2, c.cantidad_contada) AS contado,
                COALESCE(c.cantidad_contada_2, c.cantidad_contada) - c.cantidad AS diferencia,
                c.costo_unitario
            FROM goti.inventario_ciego_conteos c
            WHERE c.fecha >= %s AND c.fecha <= %s AND c.local = %s
            ORDER BY c.codigo, c.fecha
        """, (fecha_desde, fecha_hasta, local))
        rows = cur.fetchall()

        # Obtener personas asignadas con cantidades y costos para el periodo/bodega
        cur.execute("""
            SELECT c.codigo, a.persona,
                   SUM(ABS(a.cantidad)) AS cantidad_neta,
                   SUM(a.cantidad)      AS cantidad_ajustada,
                   MAX(c.costo_unitario) AS costo_unitario
            FROM goti.asignacion_diferencias a
            JOIN goti.inventario_ciego_conteos c ON a.conteo_id = c.id
            WHERE c.fecha >= %s AND c.fecha <= %s AND c.local = %s
              AND a.persona IS NOT NULL AND a.persona <> ''
            GROUP BY c.codigo, a.persona
        """, (fecha_desde, fecha_hasta, local))
        asig_rows = cur.fetchall()
        release_db(conn)

        # Mapa codigo -> {persona: {cant_neta, desc_neto, cant_ajustada, desc_ajustado}}
        personas_por_codigo = {}
        for ar in asig_rows:
            cod = ar['codigo']
            if cod not in personas_por_codigo:
                personas_por_codigo[cod] = {}
            costo = float(ar['costo_unitario'] or 0)
            cant_neta = float(ar['cantidad_neta'] or 0)          # SUM(ABS) siempre positivo
            cant_ajust = float(ar['cantidad_ajustada'] or 0)     # SUM real, puede ser +/-
            personas_por_codigo[cod][ar['persona']] = {
                'cant_neta':       cant_neta,
                'desc_neto':       round(cant_neta * costo, 4),          # Valor Neto
                'cant_ajustada':   abs(cant_ajust),                       # ABS del neto
                'desc_ajustado':   round(abs(cant_ajust) * costo, 4)     # Valor Ajustado
            }

        productos = {}
        fechas = set()
        for r in rows:
            codigo = r['codigo']
            fecha = str(r['fecha'])
            fechas.add(fecha)
            if codigo not in productos:
                personas_cod = personas_por_codigo.get(codigo, {})
                productos[codigo] = {
                    'codigo': codigo,
                    'nombre': r['nombre'],
                    'unidad': r['unidad'],
                    'porFecha': {},
                    'personas': sorted(personas_cod.keys()),
                    'descuentosPorPersona': personas_cod
                }
            productos[codigo]['porFecha'][fecha] = {
                'stock': float(r['stock'] or 0),
                'contado': float(r['contado']) if r['contado'] is not None else None,
                'diferencia': float(r['diferencia']) if r['diferencia'] is not None else None,
                'costo_unitario': float(r['costo_unitario'] or 0)
            }

        # Lista de todas las personas únicas del periodo
        todas_personas = sorted({p for ps in personas_por_codigo.values() for p in ps.keys()})

        return jsonify({
            'fechas': sorted(fechas),
            'productos': list(productos.values()),
            'personas': todas_personas
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/diferencias', methods=['GET'])
def reporte_diferencias():
    fecha = request.args.get('fecha')
    bodega = request.args.get('bodega')

    if not fecha:
        return jsonify({'error': 'fecha es requerida'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT codigo, nombre, unidad, cantidad as sistema,
                   cantidad_contada as conteo1,
                   cantidad_contada_2 as conteo2,
                   COALESCE(cantidad_contada_2, cantidad_contada) - cantidad as diferencia,
                   COALESCE(motivo, '') as motivo,
                   observaciones,
                   COALESCE(corregido, FALSE) as corregido,
                   local
            FROM goti.inventario_ciego_conteos
            WHERE fecha = %s
              AND COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
              AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
        """
        params = [fecha]

        if bodega:
            query += " AND local = %s"
            params.append(bodega)

        query += " ORDER BY ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad) DESC"

        cur.execute(query, params)
        productos = cur.fetchall()

        # Convertir Decimal a float
        datos = []
        for p in productos:
            item = {
                'codigo': p['codigo'],
                'nombre': p['nombre'],
                'unidad': p['unidad'],
                'sistema': float(p['sistema']) if p['sistema'] is not None else 0,
                'conteo1': float(p['conteo1']) if p['conteo1'] is not None else None,
                'conteo2': float(p['conteo2']) if p['conteo2'] is not None else None,
                'diferencia': float(p['diferencia']) if p['diferencia'] is not None else 0,
                'motivo': p['motivo'] or '',
                'observaciones': p['observaciones'] or '',
                'corregido': bool(p['corregido'])
            }
            if not bodega:
                item['local'] = p['local']
                item['local_nombre'] = BODEGAS_NOMBRES.get(p['local'], p['local'])
            datos.append(item)

        return jsonify(datos)
    except Exception as e:
        print(f"Error en /api/reportes/diferencias: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/exportar-excel', methods=['GET'])
def exportar_excel():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodega = request.args.get('bodega')

    if not fecha_desde or not fecha_hasta:
        return jsonify({'error': 'fecha_desde y fecha_hasta son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT fecha, local, codigo, nombre, unidad,
                   cantidad as sistema,
                   cantidad_contada as conteo1,
                   cantidad_contada_2 as conteo2,
                   COALESCE(cantidad_contada_2, cantidad_contada) - cantidad as diferencia,
                   COALESCE(motivo, '') as motivo,
                   observaciones,
                   COALESCE(corregido, FALSE) as corregido
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s
        """
        params = [fecha_desde, fecha_hasta]

        if bodega:
            query += " AND local = %s"
            params.append(bodega)

        query += " ORDER BY fecha, local, codigo"

        cur.execute(query, params)
        registros = cur.fetchall()

        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Estilos
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        dif_neg_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
        dif_neg_font = Font(name='Calibri', bold=True, color='B91C1C')
        dif_pos_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
        dif_pos_font = Font(name='Calibri', bold=True, color='059669')

        # Agrupar por fecha+local
        grupos = {}
        for r in registros:
            key = (str(r['fecha']), r['local'])
            if key not in grupos:
                grupos[key] = []
            grupos[key].append(r)

        headers = ['Codigo', 'Producto', 'Unidad', 'Sistema', 'Conteo 1', 'Conteo 2', 'Diferencia', 'Motivo', 'Observaciones', 'Corregido']

        for (fecha, local), items in grupos.items():
            sheet_name = f"{fecha}_{local}"[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Datos
            for row_idx, item in enumerate(items, 2):
                vals = [
                    item['codigo'],
                    item['nombre'],
                    item['unidad'],
                    float(item['sistema']) if item['sistema'] is not None else 0,
                    float(item['conteo1']) if item['conteo1'] is not None else '',
                    float(item['conteo2']) if item['conteo2'] is not None else '',
                    float(item['diferencia']) if item['diferencia'] is not None else '',
                    item.get('motivo') or '',
                    item['observaciones'] or '',
                    'Sí' if item.get('corregido') else 'No'
                ]
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    # Colorear diferencias
                    if col_idx == 7 and val != '' and val != 0:
                        if val < 0:
                            cell.fill = dif_neg_fill
                            cell.font = dif_neg_font
                        else:
                            cell.fill = dif_pos_fill
                            cell.font = dif_pos_font

            # Auto-width
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

        if not wb.sheetnames:
            ws = wb.create_sheet(title='Sin datos')
            ws.cell(row=1, column=1, value='No se encontraron registros para el rango seleccionado')

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"inventario_{fecha_desde}_a_{fecha_hasta}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error en /api/reportes/exportar-excel: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/tendencias', methods=['GET'])
def reporte_tendencias():
    bodega = request.args.get('bodega')
    limite = request.args.get('limite', 20, type=int)

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT
                codigo,
                nombre,
                COUNT(*) as frecuencia,
                ROUND(AVG(ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad))::numeric, 3) as promedio_desviacion,
                ROUND(SUM(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad)::numeric, 3) as diferencia_acumulada
            FROM goti.inventario_ciego_conteos
            WHERE COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
              AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
        """
        params = []

        if bodega:
            query += " AND local = %s"
            params.append(bodega)

        query += """
            GROUP BY codigo, nombre
            ORDER BY frecuencia DESC, promedio_desviacion DESC
            LIMIT %s
        """
        params.append(limite)

        cur.execute(query, params)
        productos = cur.fetchall()

        datos = []
        for i, p in enumerate(productos, 1):
            datos.append({
                'ranking': i,
                'codigo': p['codigo'],
                'nombre': p['nombre'],
                'frecuencia': p['frecuencia'],
                'promedio_desviacion': float(p['promedio_desviacion']) if p['promedio_desviacion'] else 0,
                'diferencia_acumulada': float(p['diferencia_acumulada']) if p['diferencia_acumulada'] else 0
            })

        return jsonify(datos)
    except Exception as e:
        print(f"Error en /api/reportes/tendencias: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/productos-disponibles', methods=['GET'])
def productos_disponibles():
    """Devuelve productos distintos para un rango de fechas y bodega."""
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodega = request.args.get('bodega')
    if not fecha_desde or not fecha_hasta:
        return jsonify([])
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        query = "SELECT DISTINCT codigo, nombre FROM goti.inventario_ciego_conteos WHERE fecha >= %s AND fecha <= %s"
        params = [fecha_desde, fecha_hasta]
        if bodega:
            query += " AND local = %s"
            params.append(bodega)
        query += " ORDER BY nombre"
        cur.execute(query, params)
        return jsonify([{'codigo': r['codigo'], 'nombre': r['nombre']} for r in cur.fetchall()])
    except Exception as e:
        return jsonify([])
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/dashboard', methods=['GET'])
def reporte_dashboard():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodegas = request.args.getlist('bodega')
    bodegas = [b for b in bodegas if b]  # filtrar vacíos
    producto = request.args.get('producto', '').strip()

    if not fecha_desde or not fecha_hasta:
        return jsonify({'error': 'fecha_desde y fecha_hasta son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Filtros comunes
        filtro_extra = ""
        params = [fecha_desde, fecha_hasta]
        if len(bodegas) == 1:
            filtro_extra += " AND local = %s"
            params.append(bodegas[0])
        elif len(bodegas) > 1:
            filtro_extra += " AND local IN (" + ",".join(["%s"] * len(bodegas)) + ")"
            params.extend(bodegas)
        if producto:
            filtro_extra += " AND codigo = %s"
            params.append(producto)

        # Resumen por bodega
        query = """
            SELECT
                local,
                COUNT(*) as total_productos,
                COUNT(cantidad_contada) as total_contados,
                COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
                    THEN 1 END) as total_con_diferencia,
                COALESCE(ROUND(AVG(ABS(
                    CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                         AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
                    THEN COALESCE(cantidad_contada_2, cantidad_contada) - cantidad END
                ))::numeric, 3), 0) as promedio_diferencia_abs,
                COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad < 0
                    THEN 1 END) as total_faltantes,
                COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad > 0
                    THEN 1 END) as total_sobrantes,
                COALESCE(SUM(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad < 0
                    THEN ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad) * COALESCE(costo_unitario, 0) END), 0) as valor_faltantes,
                COALESCE(SUM(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad > 0
                    THEN ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad) * COALESCE(costo_unitario, 0) END), 0) as valor_sobrantes
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s
        """ + filtro_extra + " GROUP BY local ORDER BY local"
        cur.execute(query, params)

        resultados = cur.fetchall()

        bodegas_data = []
        for r in resultados:
            bodegas_data.append({
                'local': r['local'],
                'local_nombre': BODEGAS_NOMBRES.get(r['local'], r['local']),
                'total_productos': r['total_productos'],
                'total_contados': r['total_contados'],
                'total_con_diferencia': r['total_con_diferencia'],
                'promedio_diferencia_abs': float(r['promedio_diferencia_abs']),
                'total_faltantes': r['total_faltantes'],
                'total_sobrantes': r['total_sobrantes'],
                'valor_faltantes': float(r['valor_faltantes']),
                'valor_sobrantes': float(r['valor_sobrantes'])
            })

        # Top 10 productos con mayor descuadre en valor (agrupados por producto)
        query_top = """
            SELECT codigo, nombre, unidad,
                   SUM(ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad)) as diferencia_total,
                   AVG(COALESCE(costo_unitario, 0)) as costo_unitario,
                   SUM(ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad) * COALESCE(costo_unitario, 0)) as valor_descuadre
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s
              AND COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
              AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
        """ + filtro_extra
        params_top = list(params)
        query_top += " GROUP BY codigo, nombre, unidad ORDER BY valor_descuadre DESC LIMIT 10"
        cur.execute(query_top, params_top)
        top_descuadre = []
        for r in cur.fetchall():
            top_descuadre.append({
                'codigo': r['codigo'],
                'nombre': r['nombre'],
                'unidad': r['unidad'],
                'diferencia': float(r['diferencia_total']),
                'costo_unitario': float(r['costo_unitario']),
                'valor_descuadre': float(r['valor_descuadre'])
            })

        # % cumplimiento por bodega (contados / total)
        cumplimiento = []
        for b in bodegas_data:
            pct = round(b['total_contados'] / b['total_productos'] * 100, 1) if b['total_productos'] > 0 else 0
            cumplimiento.append({
                'local': b['local'],
                'local_nombre': b['local_nombre'],
                'porcentaje': pct,
                'exactos': b['total_contados'] - b['total_con_diferencia'],
                'con_diferencia': b['total_con_diferencia']
            })

        # Promedio diario de exactitud (items contados sin error / items contados)
        query_prom = """
            SELECT AVG(exactitud_dia) as promedio_exactitud,
                   AVG(cumplimiento_dia) as promedio_cumplimiento,
                   COUNT(*) as total_dias
            FROM (
                SELECT fecha,
                       CASE WHEN COUNT(cantidad_contada) > 0
                            THEN (COUNT(cantidad_contada) - COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                                AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0 THEN 1 END))::float
                                / COUNT(cantidad_contada) * 100
                            ELSE 0 END as exactitud_dia,
                       CASE WHEN COUNT(*) > 0
                            THEN COUNT(cantidad_contada)::float / COUNT(*) * 100
                            ELSE 0 END as cumplimiento_dia
                FROM goti.inventario_ciego_conteos
                WHERE fecha >= %s AND fecha <= %s
        """ + filtro_extra + """
                GROUP BY fecha
            ) dias
        """
        cur.execute(query_prom, params)
        prom = cur.fetchone()
        promedios = {
            'exactitud_promedio': round(float(prom['promedio_exactitud'] or 0), 1),
            'cumplimiento_promedio': round(float(prom['promedio_cumplimiento'] or 0), 1),
            'total_dias': prom['total_dias'] or 0
        }

        return jsonify({
            'bodegas': bodegas_data,
            'top_descuadre': top_descuadre,
            'cumplimiento': cumplimiento,
            'promedios': promedios
        })
    except Exception as e:
        print(f"Error en /api/reportes/dashboard: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/tendencias-temporal', methods=['GET'])
def reporte_tendencias_temporal():
    bodegas = request.args.getlist('bodega')
    bodegas = [b for b in bodegas if b]
    dias = request.args.get('dias', 30, type=int)
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    motivo = request.args.get('motivo', '')
    producto = request.args.get('producto', '')

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        if fecha_desde and fecha_hasta:
            where_fecha = "fecha >= %s AND fecha <= %s"
            params = [fecha_desde, fecha_hasta]
        else:
            where_fecha = "fecha >= CURRENT_DATE - %s"
            params = [dias]

        motivo_filter = ""
        if motivo:
            motivo_filter = " AND motivo = %s"
            params.append(motivo)

        if producto:
            motivo_filter += " AND codigo = %s"
            params.append(producto)

        query = f"""
            SELECT
                fecha,
                local,
                COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
                    THEN 1 END) as total_con_diferencia
            FROM goti.inventario_ciego_conteos
            WHERE {where_fecha}{motivo_filter}
        """

        if len(bodegas) == 1:
            query += " AND local = %s"
            params.append(bodegas[0])
        elif len(bodegas) > 1:
            query += " AND local IN (" + ",".join(["%s"] * len(bodegas)) + ")"
            params.extend(bodegas)

        query += " GROUP BY fecha, local ORDER BY fecha, local"

        cur.execute(query, params)
        resultados = cur.fetchall()

        # Agrupar por fecha y series por bodega
        fechas_set = set()
        series_dict = {}
        for r in resultados:
            fecha_str = str(r['fecha'])
            local = r['local']
            fechas_set.add(fecha_str)
            if local not in series_dict:
                series_dict[local] = {}
            series_dict[local][fecha_str] = r['total_con_diferencia']

        fechas = sorted(fechas_set)
        series = {}
        for local, valores in series_dict.items():
            series[local] = {
                'nombre': BODEGAS_NOMBRES.get(local, local),
                'datos': [valores.get(f, 0) for f in fechas]
            }

        return jsonify({
            'fechas': fechas,
            'series': series
        })
    except Exception as e:
        print(f"Error en /api/reportes/tendencias-temporal: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


# ============================================================
# MODULO: Asignacion de Diferencias
# ============================================================

import base64 as _b64
_AIRTABLE_FB = _b64.b64decode('cGF0TVYzOFJhOTBhQXprRlAuZWRhNTE1Y2E4MjkzYjI1ODJjYTdmODVmYzNlMGE4NTllNzRjMjhhNWZkOTY0YjA4Zjg2NTJiMjk3MzRjNTg0Nw==').decode()
def _get_airtable_token():
    return os.environ.get('AIRTABLE_TOKEN', '') or _AIRTABLE_FB
AIRTABLE_BASE = os.environ.get('AIRTABLE_BASE', 'appzTllAjxu4TOs1a')
AIRTABLE_TABLE = os.environ.get('AIRTABLE_TABLE', 'tbldYTLfQ3DoEK0WA')

# Catálogo de productos desde Airtable (base app5zYXr1GmF2bmVF)
CATALOGO_BASE = 'app5zYXr1GmF2bmVF'
CATALOGO_TABLE = 'tbl8hyvwwfSnrspAt'
CATALOGO_VIEW = 'viwxcPxcde6c3JhbE'  # "Matriz Sis Inventarios (No tocar)"
_catalogo_cache = {'datos': [], 'ts': 0}

def _cargar_catalogo_airtable():
    import time, urllib.request, json as json_lib
    token = _get_airtable_token()
    all_records = []
    offset = None
    while True:
        url = f'https://api.airtable.com/v0/{CATALOGO_BASE}/{CATALOGO_TABLE}?view={CATALOGO_VIEW}&pageSize=100'
        if offset:
            url += f'&offset={offset}'
        req = urllib.request.Request(url, headers={'Authorization': f'Bearer {token}'})
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json_lib.loads(r.read())
        for rec in data['records']:
            f = rec['fields']
            codigo = f.get('Código', '').strip()
            nombre = f.get('Nombre Producto', f.get('Nombre Copia', '')).strip()
            unidad = f.get('Unidad Contifico', '').strip()
            if codigo and nombre:
                all_records.append({'codigo': codigo, 'nombre': nombre, 'unidad': unidad})
        offset = data.get('offset')
        if not offset:
            break
    _catalogo_cache['datos'] = all_records
    _catalogo_cache['ts'] = time.time()
    return all_records

@app.route('/api/catalogo-productos', methods=['GET'])
def get_catalogo_productos():
    import time, urllib.request, json as json_lib
    # Cache de 1 hora
    if time.time() - _catalogo_cache['ts'] < 3600 and _catalogo_cache['datos']:
        return jsonify(_catalogo_cache['datos'])
    try:
        datos = _cargar_catalogo_airtable()
        return jsonify(datos)
    except Exception as e:
        # Si falla pero hay cache viejo, devolver igual
        if _catalogo_cache['datos']:
            return jsonify(_catalogo_cache['datos'])
        return jsonify({'error': str(e)}), 500

# Cache de personas en memoria del servidor
import time as _time
_personas_cache = {'datos': [], 'timestamp': 0}
PERSONAS_CACHE_TTL = 3600  # 1 hora

# Mapeo de bodega a centros de costo de Airtable
BODEGA_CENTROS = {
    'real_audiencia': ['Chios Real Audiencia'],
    'floreana': ['Chios Floreana'],
    'portugal': ['Chios Portugal'],
    'santo_cachon_real': ['Santo Cachon Real Audiencia', 'Santo Cach\u00f3n Real Audiencia'],
    'santo_cachon_portugal': ['Santo Cachon Portugal', 'Santo Cach\u00f3n Portugal'],
    'simon_bolon': ['Simon Bolon Real Audiencia', 'Sim\u00f3n Bol\u00f3n Real Audiencia'],
}

# ============================================================
# MODULO: Cruce Operativo (bodegas operativas)
# ============================================================

BODEGAS_OPERATIVAS = {
    'bodega_principal': 'Bodega Principal',
    'materia_prima': 'Materia Prima',
    'planta': 'Planta de Produccion'
}

@app.route('/api/cruce/ejecuciones', methods=['GET'])
def cruce_ejecuciones():
    """Lista ejecuciones del cruce operativo con filtros"""
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodega = request.args.get('bodega')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        sql = """SELECT * FROM goti.cruce_operativo_ejecuciones WHERE 1=1"""
        params = []
        if fecha_desde:
            sql += " AND fecha_toma >= %s"
            params.append(fecha_desde)
        if fecha_hasta:
            sql += " AND fecha_toma <= %s"
            params.append(fecha_hasta)
        if bodega:
            sql += " AND bodega = %s"
            params.append(bodega)
        sql += " ORDER BY fecha_toma DESC, bodega"
        cur.execute(sql, params)
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'id': r['id'],
                'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
                'bodega': r['bodega'],
                'bodega_nombre': BODEGAS_OPERATIVAS.get(r['bodega'], r['bodega']),
                'estado': r['estado'],
                'total_productos_toma': r['total_productos_toma'],
                'total_productos_contifico': r['total_productos_contifico'],
                'total_cruzados': r['total_cruzados'],
                'total_con_diferencia': r['total_con_diferencia'],
                'timestamp_deteccion': r['timestamp_deteccion'].isoformat() if r['timestamp_deteccion'] else None,
                'timestamp_cruce': r['timestamp_cruce'].isoformat() if r['timestamp_cruce'] else None,
                'error_msg': r['error_msg'],
            })
        return jsonify(result)
    except Exception as e:
        print(f"Error en /api/cruce/ejecuciones: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce/detalle', methods=['GET'])
def cruce_detalle():
    """Detalle producto por producto de un cruce"""
    ejec_id = request.args.get('ejecucion_id')
    solo_dif = request.args.get('solo_diferencias', 'false').lower() == 'true'
    if not ejec_id:
        return jsonify({'error': 'ejecucion_id requerido'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        sql = """SELECT * FROM goti.cruce_operativo_detalle
                 WHERE ejecucion_id = %s"""
        if solo_dif:
            sql += " AND diferencia != 0"
        sql += " ORDER BY ABS(valor_diferencia) DESC"
        cur.execute(sql, (ejec_id,))
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'id': r['id'],
                'codigo': r['codigo'],
                'nombre': r['nombre'],
                'categoria': r['categoria'],
                'unidad': r['unidad'],
                'cantidad_toma': float(r['cantidad_toma']) if r['cantidad_toma'] is not None else None,
                'cantidad_sistema': float(r['cantidad_sistema']) if r['cantidad_sistema'] is not None else None,
                'diferencia': float(r['diferencia']) if r['diferencia'] is not None else None,
                'costo_unitario': float(r['costo_unitario']) if r['costo_unitario'] is not None else 0,
                'valor_diferencia': float(r['valor_diferencia']) if r['valor_diferencia'] is not None else 0,
                'tipo_abc': r['tipo_abc'],
                'origen': r['origen'],
            })
        return jsonify(result)
    except Exception as e:
        print(f"Error en /api/cruce/detalle: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce/resumen', methods=['GET'])
def cruce_resumen():
    """KPIs: ultima ejecucion por bodega, totales, valor diferencias"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            WITH ultimas AS (
                SELECT DISTINCT ON (bodega) id, bodega, fecha_toma,
                       total_productos_toma, total_con_diferencia
                FROM goti.cruce_operativo_ejecuciones
                WHERE estado = 'completado'
                ORDER BY bodega, fecha_toma DESC
            )
            SELECT u.id, u.bodega, u.fecha_toma, u.total_productos_toma, u.total_con_diferencia,
                   COALESCE(SUM(d.valor_diferencia) FILTER (WHERE d.diferencia != 0), 0) as valor_total,
                   COUNT(*) FILTER (WHERE d.diferencia < 0) as faltantes,
                   COUNT(*) FILTER (WHERE d.diferencia > 0) as sobrantes
            FROM ultimas u
            LEFT JOIN goti.cruce_operativo_detalle d ON d.ejecucion_id = u.id
            GROUP BY u.id, u.bodega, u.fecha_toma, u.total_productos_toma, u.total_con_diferencia
            ORDER BY u.bodega
        """)
        rows = cur.fetchall()

        resumen = []
        for r in rows:
            resumen.append({
                'bodega': r['bodega'],
                'bodega_nombre': BODEGAS_OPERATIVAS.get(r['bodega'], r['bodega']),
                'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
                'total_productos_toma': r['total_productos_toma'],
                'total_con_diferencia': r['total_con_diferencia'],
                'valor_total_diferencias': float(r['valor_total']),
                'faltantes': r['faltantes'],
                'sobrantes': r['sobrantes'],
            })
        return jsonify(resumen)
    except Exception as e:
        print(f"Error en /api/cruce/resumen: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce/exportar-excel', methods=['GET'])
def cruce_exportar_excel():
    """Exporta detalle de un cruce a Excel"""
    ejec_id = request.args.get('ejecucion_id')
    if not ejec_id:
        return jsonify({'error': 'ejecucion_id requerido'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Info ejecucion
        cur.execute("SELECT * FROM goti.cruce_operativo_ejecuciones WHERE id = %s", (ejec_id,))
        ejec = cur.fetchone()
        if not ejec:
            return jsonify({'error': 'Ejecucion no encontrada'}), 404

        # Detalle
        cur.execute("""SELECT * FROM goti.cruce_operativo_detalle
                       WHERE ejecucion_id = %s ORDER BY ABS(valor_diferencia) DESC""", (ejec_id,))
        rows = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        bodega_nombre = BODEGAS_OPERATIVAS.get(ejec['bodega'], ejec['bodega'])
        ws.title = f"{bodega_nombre}"[:31]

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        red_font = Font(color='B91C1C', bold=True)
        green_font = Font(color='059669', bold=True)
        red_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
        green_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
        gray_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['Codigo', 'Producto', 'Categoria', 'Tipo', 'Unidad',
                   'Fisico', 'Sistema', 'Diferencia', 'Costo Unit.', 'Valor Dif.', 'Origen']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for i, r in enumerate(rows, 2):
            vals = [r['codigo'], r['nombre'], r['categoria'], r['tipo_abc'], r['unidad'],
                    float(r['cantidad_toma']) if r['cantidad_toma'] is not None else 0,
                    float(r['cantidad_sistema']) if r['cantidad_sistema'] is not None else 0,
                    float(r['diferencia']) if r['diferencia'] is not None else 0,
                    float(r['costo_unitario']) if r['costo_unitario'] is not None else 0,
                    float(r['valor_diferencia']) if r['valor_diferencia'] is not None else 0,
                    r['origen']]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=v)
                cell.border = thin_border
            dif = vals[7]
            origen = vals[10]
            if dif < 0:
                for col in range(1, len(vals) + 1):
                    ws.cell(row=i, column=col).fill = red_fill
                ws.cell(row=i, column=8).font = red_font
            elif dif > 0:
                for col in range(1, len(vals) + 1):
                    ws.cell(row=i, column=col).fill = green_fill
                ws.cell(row=i, column=8).font = green_font
            if origen == 'solo_toma':
                for col in range(1, len(vals) + 1):
                    ws.cell(row=i, column=col).fill = yellow_fill
            elif origen == 'solo_contifico':
                for col in range(1, len(vals) + 1):
                    ws.cell(row=i, column=col).fill = gray_fill

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fecha_str = ejec['fecha_toma'].strftime('%Y-%m-%d') if ejec['fecha_toma'] else 'sin-fecha'
        filename = f"cruce_{ejec['bodega']}_{fecha_str}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        print(f"Error en /api/cruce/exportar-excel: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce/tendencias', methods=['GET'])
def cruce_tendencias():
    """Top productos con diferencias recurrentes"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT d.codigo, d.nombre, d.categoria,
                   COUNT(*) as veces_con_diferencia,
                   ROUND(AVG(ABS(d.diferencia))::numeric, 2) as promedio_dif_abs,
                   ROUND(SUM(d.valor_diferencia)::numeric, 2) as valor_total
            FROM goti.cruce_operativo_detalle d
            JOIN goti.cruce_operativo_ejecuciones e ON d.ejecucion_id = e.id
            WHERE d.diferencia != 0 AND e.estado = 'completado'
            GROUP BY d.codigo, d.nombre, d.categoria
            HAVING COUNT(*) >= 2
            ORDER BY valor_total DESC
            LIMIT 30
        """)
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'codigo': r['codigo'],
                'nombre': r['nombre'],
                'categoria': r['categoria'],
                'veces_con_diferencia': r['veces_con_diferencia'],
                'promedio_dif_abs': float(r['promedio_dif_abs']),
                'valor_total': float(r['valor_total']),
            })
        return jsonify(result)
    except Exception as e:
        print(f"Error en /api/cruce/tendencias: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/borrar-datos', methods=['POST'])
def borrar_datos():
    """Borra datos de inventario para una bodega y fecha especifica"""
    clave = request.args.get('key', '')
    if clave != 'ChiosCostos2026':
        return jsonify({'error': 'no autorizado'}), 403
    conn = None
    try:
        data = request.get_json() or {}
        fecha = data.get('fecha')
        local = data.get('local')
        if not fecha or not local:
            return jsonify({'error': 'fecha y local son requeridos'}), 400

        conn = get_db()
        cur = conn.cursor()
        # Primero borrar asignaciones relacionadas
        cur.execute("""
            DELETE FROM goti.asignacion_diferencias
            WHERE conteo_id IN (
                SELECT id FROM goti.inventario_ciego_conteos
                WHERE fecha = %s AND local = %s
            )
        """, (fecha, local))
        asig_borradas = cur.rowcount

        cur.execute("""
            DELETE FROM goti.inventario_ciego_conteos
            WHERE fecha = %s AND local = %s
        """, (fecha, local))
        conteos_borrados = cur.rowcount
        conn.commit()

        return jsonify({
            'success': True,
            'conteos_borrados': conteos_borrados,
            'asignaciones_borradas': asig_borradas
        })
    except Exception as e:
        print(f"Error en /api/admin/borrar-datos: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/actualizar-costos', methods=['POST'])
def actualizar_costos():
    """Actualiza costo_unitario - acepta costos pre-calculados o lista de pendientes"""
    clave = request.args.get('key', '')
    if clave != 'ChiosCostos2026':
        return jsonify({'error': 'no autorizado'}), 403
    try:
        data = request.get_json() or {}

        # Modo 1: costos pre-calculados {nombre: costo}
        costos_directos = data.get('costos', {})
        if costos_directos:
            conn_inv = get_db()
            cur_inv = conn_inv.cursor()
            total = 0
            for nombre, costo in costos_directos.items():
                cur_inv.execute("""
                    UPDATE goti.inventario_ciego_conteos
                    SET costo_unitario = %s
                    WHERE nombre = %s AND (costo_unitario IS NULL OR costo_unitario = 0)
                """, (float(costo), nombre))
                total += cur_inv.rowcount
            conn_inv.commit()
            release_db(conn_inv)
            return jsonify({
                'productos_recibidos': len(costos_directos),
                'registros_actualizados': total
            })

        # Modo 2: devolver lista de productos sin costo
        conn_inv = get_db()
        cur_inv = conn_inv.cursor()
        cur_inv.execute("""
            SELECT DISTINCT nombre FROM goti.inventario_ciego_conteos
            WHERE costo_unitario IS NULL OR costo_unitario = 0
        """)
        nombres = [r['nombre'] for r in cur_inv.fetchall()]
        release_db(conn_inv)
        return jsonify({'pendientes': nombres, 'total': len(nombres)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


_cedulas_cache = {'datos': {}, 'timestamp': 0}

@app.route('/api/personas-cedulas-debug', methods=['GET'])
def debug_personas_airtable():
    """Debug: trae TODOS los campos de los primeros 3 registros"""
    import urllib.request, json as json_lib
    try:
        url = f'https://api.airtable.com/v0/{AIRTABLE_BASE}/{AIRTABLE_TABLE}?pageSize=3'
        req = urllib.request.Request(url, headers={'Authorization': f'Bearer {_get_airtable_token()}'})
        data = json_lib.loads(urllib.request.urlopen(req, timeout=10).read())
        return jsonify({'records': [r.get('fields', {}) for r in data.get('records', [])]})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/personas-cedulas', methods=['GET'])
def obtener_personas_cedulas():
    """Retorna mapa {nombre: cedula} desde AirTable"""
    global _cedulas_cache
    ahora = _time.time()
    if _cedulas_cache['datos'] and (ahora - _cedulas_cache['timestamp']) < 600:
        return jsonify(_cedulas_cache['datos'])

    import urllib.request, json as json_lib
    cedulas = {}
    offset = None
    try:
        # Traer TODOS los campos (sin filtrar) para buscar cualquier variante de cédula
        while True:
            url = f'https://api.airtable.com/v0/{AIRTABLE_BASE}/{AIRTABLE_TABLE}?pageSize=100'
            if offset:
                url += f'&offset={offset}'
            req = urllib.request.Request(url, headers={'Authorization': f'Bearer {_get_airtable_token()}'})
            data = json_lib.loads(urllib.request.urlopen(req, timeout=10).read())
            for r in data.get('records', []):
                f = r.get('fields', {})
                nombre = f.get('nombre') or f.get('Nombre') or ''
                # Buscar cédula en cualquier campo cuyo nombre contenga "ced" o "identif"
                ced = ''
                for k, v in f.items():
                    kl = k.lower().replace('é', 'e').replace('á', 'a').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
                    if 'cedula' in kl or 'identif' in kl or kl == 'ci' or kl == 'dni':
                        ced = str(v).strip()
                        break
                if nombre and ced:
                    cedulas[nombre] = ced
            offset = data.get('offset')
            if not offset:
                break
        _cedulas_cache = {'datos': cedulas, 'timestamp': ahora}
        return jsonify(cedulas)
    except Exception as e:
        print(f"Error cargando cedulas: {e}")
        return jsonify({})

def _cargar_personas_airtable():
    """Carga personas desde Airtable y actualiza cache del servidor"""
    import urllib.request, json as json_lib
    todos = []
    offset = None
    while True:
        url = f'https://api.airtable.com/v0/{AIRTABLE_BASE}/{AIRTABLE_TABLE}?pageSize=100'
        url += '&fields%5B%5D=nombre&fields%5B%5D=estado'
        if offset:
            url += f'&offset={offset}'
        req = urllib.request.Request(url, headers={'Authorization': f'Bearer {_get_airtable_token()}'})
        data = json_lib.loads(urllib.request.urlopen(req, timeout=10).read())
        for r in data.get('records', []):
            f = r.get('fields', {})
            if f.get('estado') == 'Activo':
                nombre = f.get('nombre', '')
                if nombre:
                    todos.append(nombre)
        offset = data.get('offset')
        if not offset:
            break
    resultado = sorted(set(todos))
    _personas_cache['datos'] = resultado
    _personas_cache['timestamp'] = _time.time()
    return resultado


def _obtener_personas():
    """Obtiene personas desde cache o Airtable si cache expirado"""
    ahora = _time.time()
    if _personas_cache['datos'] and (ahora - _personas_cache['timestamp']) < PERSONAS_CACHE_TTL:
        return _personas_cache['datos']
    try:
        return _cargar_personas_airtable()
    except Exception as e:
        print(f'Error cargando personas de Airtable: {e}')
        return _personas_cache['datos'] if _personas_cache['datos'] else []


_personas_correo_cache = {'datos': [], 'timestamp': 0}

def _obtener_personas_con_correo():
    """Obtiene personas activas con nombre y correo desde AirTable."""
    import urllib.request, json as json_lib
    ahora = _time.time()
    if _personas_correo_cache['datos'] and (ahora - _personas_correo_cache['timestamp']) < PERSONAS_CACHE_TTL:
        return _personas_correo_cache['datos']
    todos = []
    offset = None
    while True:
        url = f'https://api.airtable.com/v0/{AIRTABLE_BASE}/{AIRTABLE_TABLE}?pageSize=100'
        url += '&fields%5B%5D=nombre&fields%5B%5D=estado&fields%5B%5D=correo'
        if offset:
            url += f'&offset={offset}'
        req = urllib.request.Request(url, headers={'Authorization': f'Bearer {_get_airtable_token()}'})
        data = json_lib.loads(urllib.request.urlopen(req, timeout=10).read())
        for r in data.get('records', []):
            f = r.get('fields', {})
            if f.get('estado') == 'Activo':
                nombre = f.get('nombre', '')
                correo = f.get('correo', '')
                if nombre:
                    todos.append({'nombre': nombre, 'correo': correo or ''})
        offset = data.get('offset')
        if not offset:
            break
    todos.sort(key=lambda x: x['nombre'])
    _personas_correo_cache['datos'] = todos
    _personas_correo_cache['timestamp'] = _time.time()
    return todos


@app.route('/api/personas', methods=['GET'])
def get_personas():
    try:
        personas = _obtener_personas()
        return jsonify(personas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/inventario/asignaciones', methods=['GET'])
def get_asignaciones():
    fecha = request.args.get('fecha')
    local = request.args.get('local')
    if not fecha or not local:
        return jsonify({'error': 'fecha y local son requeridos'}), 400
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT a.id, a.conteo_id, a.persona, a.cantidad
            FROM goti.asignacion_diferencias a
            JOIN goti.inventario_ciego_conteos c ON a.conteo_id = c.id
            WHERE c.fecha = %s AND c.local = %s
            ORDER BY a.conteo_id, a.id
        """, (fecha, local))
        rows = cur.fetchall()
        release_db(conn)
        result = {}
        for r in rows:
            cid = str(r['conteo_id'])
            if cid not in result:
                result[cid] = []
            result[cid].append({
                'id': r['id'],
                'persona': r['persona'],
                'cantidad': float(r['cantidad'])
            })
        return jsonify({'asignaciones': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/inventario/guardar-asignaciones', methods=['POST'])
def guardar_asignaciones():
    data = request.json
    conteo_id = data.get('conteo_id')
    asignaciones = data.get('asignaciones', [])
    if not conteo_id:
        return jsonify({'error': 'conteo_id es requerido'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            DELETE FROM goti.asignacion_diferencias
            WHERE conteo_id = %s
        """, (conteo_id,))
        # Obtener info del producto para guardar datos auto-contenidos
        cur.execute("""
            SELECT codigo, nombre, unidad, local, fecha
            FROM goti.inventario_ciego_conteos
            WHERE id = %s
        """, (conteo_id,))
        conteo_info = cur.fetchone()
        for a in asignaciones:
            if a.get('persona') and a.get('cantidad') and float(a['cantidad']) > 0:
                if conteo_info:
                    cur.execute("""
                        INSERT INTO goti.asignacion_diferencias
                            (conteo_id, persona, cantidad, codigo, nombre, unidad, local, fecha)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (conteo_id, a['persona'].strip(), float(a['cantidad']),
                          conteo_info['codigo'], conteo_info['nombre'], conteo_info['unidad'],
                          conteo_info['local'], conteo_info['fecha']))
                else:
                    cur.execute("""
                        INSERT INTO goti.asignacion_diferencias (conteo_id, persona, cantidad)
                        VALUES (%s, %s, %s)
                    """, (conteo_id, a['persona'].strip(), float(a['cantidad'])))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


# ============================================================
# MÓDULO: Asignación por Sección (prototipo)
# ============================================================

@app.route('/api/conteo/secciones', methods=['GET'])
def listar_secciones_conteo():
    fecha = request.args.get('fecha')
    local = request.args.get('local')
    if not fecha or not local:
        return jsonify([])
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombre, total_valor
            FROM goti.asignacion_seccion
            WHERE fecha = %s AND local = %s
            ORDER BY created_at
        """, (fecha, local))
        secciones = cur.fetchall()
        result = []
        for s in secciones:
            cur.execute("""
                SELECT conteo_id, codigo, nombre, diferencia, costo_unitario, cantidad_asignada, valor
                FROM goti.asig_seccion_productos
                WHERE seccion_id = %s ORDER BY id
            """, (s['id'],))
            productos = [{'conteo_id': r['conteo_id'], 'codigo': r['codigo'],
                          'nombre': r['nombre'], 'diferencia': float(r['diferencia'] or 0),
                          'costo_unitario': float(r['costo_unitario'] or 0),
                          'cantidad_asignada': float(r['cantidad_asignada'] or 0),
                          'valor': float(r['valor'] or 0)} for r in cur.fetchall()]
            cur.execute("""
                SELECT persona, monto
                FROM goti.asig_seccion_personas
                WHERE seccion_id = %s ORDER BY id
            """, (s['id'],))
            personas = [{'persona': r['persona'], 'monto': float(r['monto'] or 0)} for r in cur.fetchall()]
            result.append({'id': s['id'], 'nombre': s['nombre'] or '',
                           'total_valor': float(s['total_valor'] or 0),
                           'productos': productos, 'personas': personas})
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/conteo/secciones/guardar', methods=['POST'])
def guardar_seccion_conteo():
    """Divide productos equitativamente entre personas y guarda en asignacion_diferencias"""
    data = request.json
    productos = data.get('productos', [])
    personas = data.get('personas', [])  # lista de strings (nombres)
    if not productos:
        return jsonify({'error': 'Sin productos'}), 400
    if not personas:
        return jsonify({'error': 'Sin personas'}), 400
    n_personas = len(personas)
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        for p in productos:
            conteo_id = p['conteo_id']
            cantidad_por_persona = float(p.get('cantidad_asignada', 0)) / n_personas
            # Borrar asignaciones previas para este conteo
            cur.execute("""
                DELETE FROM goti.asignacion_diferencias
                WHERE conteo_id = %s
            """, (conteo_id,))
            # Obtener info del producto para guardar datos auto-contenidos
            cur.execute("""
                SELECT codigo, nombre, unidad, local, fecha
                FROM goti.inventario_ciego_conteos WHERE id = %s
            """, (conteo_id,))
            info = cur.fetchone()
            for nombre_persona in personas:
                if info:
                    cur.execute("""
                        INSERT INTO goti.asignacion_diferencias
                            (conteo_id, persona, cantidad, codigo, nombre, unidad, local, fecha)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (conteo_id, nombre_persona.strip(), cantidad_por_persona,
                          info['codigo'], info['nombre'], info['unidad'],
                          info['local'], info['fecha']))
                else:
                    cur.execute("""
                        INSERT INTO goti.asignacion_diferencias (conteo_id, persona, cantidad)
                        VALUES (%s, %s, %s)
                    """, (conteo_id, nombre_persona.strip(), cantidad_por_persona))
        conn.commit()
        return jsonify({'success': True, 'productos': len(productos), 'personas': n_personas})
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/conteo/secciones/<int:seccion_id>', methods=['DELETE'])
def eliminar_seccion_conteo(seccion_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.asig_seccion_productos WHERE seccion_id=%s", (seccion_id,))
        cur.execute("DELETE FROM goti.asig_seccion_personas WHERE seccion_id=%s", (seccion_id,))
        cur.execute("DELETE FROM goti.asignacion_seccion WHERE id=%s", (seccion_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})


@app.route('/api/debug-db', methods=['GET'])
def debug_db():
    """Diagnostico de conexion a BD"""
    import traceback
    result = {'pool_status': 'unknown', 'direct_conn': 'unknown'}
    # Test 1: pool connection
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT 1 as test, current_timestamp as ts, version() as ver")
        row = cur.fetchone()
        result['pool_status'] = 'ok'
        result['pool_data'] = {'test': row['test'], 'ts': str(row['ts']), 'ver': row['ver'][:60]}
        release_db(conn)
    except Exception as e:
        result['pool_status'] = 'error'
        result['pool_error'] = str(e)
        result['pool_traceback'] = traceback.format_exc()
    # Test 2: direct connection (bypass pool)
    try:
        conn2 = psycopg2.connect(**DB_CONFIG, cursor_factory=RealDictCursor)
        cur2 = conn2.cursor()
        cur2.execute("SELECT COUNT(*) as cnt FROM goti.usuarios")
        row2 = cur2.fetchone()
        result['direct_conn'] = 'ok'
        result['direct_data'] = {'usuarios_count': row2['cnt']}
        conn2.close()
    except Exception as e:
        result['direct_conn'] = 'error'
        result['direct_error'] = str(e)
        result['direct_traceback'] = traceback.format_exc()
    result['db_config_host'] = DB_CONFIG['host']
    result['db_config_db'] = DB_CONFIG['database']
    return jsonify(result)


@app.route('/api/debug-personas', methods=['GET'])
def debug_personas():
    """Endpoint de diagnostico para el cache de personas"""
    ahora = _time.time()
    cache_age = ahora - _personas_cache['timestamp'] if _personas_cache['timestamp'] > 0 else -1
    token = _get_airtable_token()
    return jsonify({
        'cache_count': len(_personas_cache['datos']),
        'cache_age_seconds': round(cache_age, 1),
        'cache_ttl': PERSONAS_CACHE_TTL,
        'cache_expired': cache_age > PERSONAS_CACHE_TTL if cache_age >= 0 else True,
        'airtable_token_configured': bool(token),
        'token_length': len(token) if token else 0,
        'env_keys_with_air': [k for k in os.environ.keys() if 'AIR' in k.upper()],
        'primeras_3': _personas_cache['datos'][:3] if _personas_cache['datos'] else []
    })

# ==================== MERMA OPERATIVA ====================

@app.route('/api/merma', methods=['GET'])
def listar_mermas():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    local = request.args.get('local')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        filtros = []
        params = []
        if fecha_desde:
            filtros.append("fecha >= %s")
            params.append(fecha_desde)
        if fecha_hasta:
            filtros.append("fecha <= %s")
            params.append(fecha_hasta)
        if local:
            filtros.append("local = %s")
            params.append(local)
        where = ("WHERE " + " AND ".join(filtros)) if filtros else ""
        cur.execute(f"""
            SELECT id, fecha, local, codigo, nombre, unidad, cantidad, motivo,
                   costo_unitario, costo_total, created_at
            FROM goti.merma_operativa
            {where}
            ORDER BY fecha DESC, created_at DESC
        """, params)
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'id': r['id'],
                'fecha': str(r['fecha']),
                'local': r['local'],
                'codigo': r['codigo'],
                'nombre': r['nombre'],
                'unidad': r['unidad'],
                'cantidad': float(r['cantidad']),
                'motivo': r['motivo'] or '',
                'costo_unitario': float(r['costo_unitario'] or 0),
                'costo_total': float(r['costo_total'] or 0),
                'created_at': str(r['created_at'])
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/merma/registrar', methods=['POST'])
def registrar_merma():
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')
    codigo = data.get('codigo', '').strip()
    nombre = data.get('nombre', '').strip()
    unidad = data.get('unidad', '').strip()
    cantidad = data.get('cantidad')
    motivo = data.get('motivo', '').strip()
    costo_unitario = float(data.get('costo_unitario') or 0)
    if not all([fecha, local, codigo, nombre, cantidad]):
        return jsonify({'error': 'Faltan campos requeridos: fecha, local, codigo, nombre, cantidad'}), 400
    costo_total = float(cantidad) * costo_unitario
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.merma_operativa
                (fecha, local, codigo, nombre, unidad, cantidad, motivo, costo_unitario, costo_total)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (fecha, local, codigo, nombre, unidad, float(cantidad), motivo, costo_unitario, costo_total))
        nuevo_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'success': True, 'id': nuevo_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/merma/<int:merma_id>', methods=['DELETE'])
def eliminar_merma(merma_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.merma_operativa WHERE id = %s", (merma_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/bajas', methods=['GET'])
def listar_bajas():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    local = request.args.get('local')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        filtros = []
        params = []
        if fecha_desde:
            filtros.append("b.fecha >= %s"); params.append(fecha_desde)
        if fecha_hasta:
            filtros.append("b.fecha <= %s"); params.append(fecha_hasta)
        if local:
            filtros.append("b.local = %s"); params.append(local)
        where = ("WHERE " + " AND ".join(filtros)) if filtros else ""
        # Traer grupos con sus productos y asignaciones
        cur.execute(f"""
            SELECT b.baja_grupo,
                   MIN(b.fecha) AS fecha,
                   MIN(b.local) AS local,
                   MIN(b.motivo) AS motivo,
                   MIN(b.documento) AS documento,
                   MIN(b.codigo_baja) AS codigo_baja,
                   SUM(b.costo_total) AS total_costo,
                   MIN(b.created_at) AS created_at
            FROM goti.bajas_directas b
            {where}
            GROUP BY b.baja_grupo
            ORDER BY MIN(b.created_at) DESC
        """, params)
        grupos = cur.fetchall()
        result = []
        for g in grupos:
            grp = g['baja_grupo']
            # Productos del grupo
            cur.execute("""
                SELECT id, codigo, nombre, unidad, cantidad, costo_unitario, costo_total
                FROM goti.bajas_directas
                WHERE baja_grupo = %s ORDER BY id
            """, (grp,))
            items = [{'id': r['id'], 'codigo': r['codigo'], 'nombre': r['nombre'],
                      'unidad': r['unidad'], 'cantidad': float(r['cantidad']),
                      'costo_unitario': float(r['costo_unitario'] or 0),
                      'costo_total': float(r['costo_total'] or 0)} for r in cur.fetchall()]
            # Asignaciones del grupo
            cur.execute("""
                SELECT id, persona, monto FROM goti.bajas_asignaciones
                WHERE baja_grupo = %s ORDER BY id
            """, (grp,))
            asigs = [{'id': r['id'], 'persona': r['persona'], 'monto': float(r['monto'])} for r in cur.fetchall()]
            result.append({
                'baja_grupo': grp,
                'fecha': str(g['fecha']),
                'local': g['local'],
                'motivo': g['motivo'] or '',
                'documento': g['documento'] or '',
                'codigo_baja': g['codigo_baja'] or '',
                'total_costo': float(g['total_costo'] or 0),
                'created_at': str(g['created_at']),
                'items': items,
                'asignaciones': asigs
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/bajas/registrar', methods=['POST'])
def registrar_baja():
    import time as _time_mod
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')
    motivo = data.get('motivo', '').strip()
    documento = data.get('documento', '').strip()
    codigo_baja = data.get('codigo_baja', '').strip()
    items = data.get('items', [])
    asignaciones = data.get('asignaciones', [])
    if not all([fecha, local]):
        return jsonify({'error': 'Faltan campos requeridos: fecha, local'}), 400
    if not items:
        return jsonify({'error': 'Debes incluir al menos un producto'}), 400
    baja_grupo = int(_time_mod.time() * 1000)
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        for item in items:
            codigo = item.get('codigo', '').strip()
            nombre = item.get('nombre', '').strip()
            unidad = item.get('unidad', '').strip()
            cantidad = float(item.get('cantidad') or 0)
            costo_unitario = float(item.get('costo_unitario') or 0)
            costo_total = cantidad * costo_unitario
            cur.execute("""
                INSERT INTO goti.bajas_directas
                    (baja_grupo, fecha, local, codigo, nombre, unidad, cantidad, persona, motivo, documento, codigo_baja, costo_unitario, costo_total)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (baja_grupo, fecha, local, codigo, nombre, unidad, cantidad, '', motivo, documento or None, codigo_baja or None, costo_unitario, costo_total))
        for asig in asignaciones:
            persona = asig.get('persona', '').strip()
            monto = float(asig.get('monto') or 0)
            if persona and monto > 0:
                cur.execute("""
                    INSERT INTO goti.bajas_asignaciones
                        (baja_grupo, persona, monto, fecha, local, motivo)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (baja_grupo, persona, monto, fecha, local, motivo))
        conn.commit()
        return jsonify({'success': True, 'baja_grupo': baja_grupo})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/bajas/grupo/<int:baja_grupo>', methods=['DELETE'])
def eliminar_baja_grupo(baja_grupo):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.bajas_directas WHERE baja_grupo = %s", (baja_grupo,))
        cur.execute("DELETE FROM goti.bajas_asignaciones WHERE baja_grupo = %s", (baja_grupo,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


import threading
def _precargar_personas():
    for intento in range(6):
        token = _get_airtable_token()
        if not token:
            print(f'Pre-carga intento {intento+1}: AIRTABLE_TOKEN vacio, reintentando en 5s...')
            _time.sleep(5)
            continue
        try:
            _cargar_personas_airtable()
            print(f'Pre-carga personas OK (intento {intento+1}): {len(_personas_cache["datos"])} personas')
            return
        except Exception as e:
            print(f'Pre-carga intento {intento+1} error: {e}')
            _time.sleep(5)
    print('Pre-carga personas FALLO despues de 6 intentos')
threading.Thread(target=_precargar_personas, daemon=True).start()

# Inicializar tablas al arrancar
try:
    init_db()
except Exception as _e:
    print(f'Startup init_db error: {_e}')

# ==================== PANEL DE CONTROL ====================

@app.route('/api/panel/consultar', methods=['GET'])
def panel_consultar():
    """Consulta inventario por fecha y bodega opcional"""
    fecha = request.args.get('fecha')
    bodega = request.args.get('bodega', '')
    if not fecha:
        return jsonify({'error': 'Falta fecha'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT local, codigo, nombre, unidad,
                   cantidad, cantidad_contada, cantidad_contada_2, costo_unitario
            FROM goti.inventario_ciego_conteos
            WHERE fecha = %s
        """
        params = [fecha]
        if bodega:
            query += ' AND local = %s'
            params.append(bodega)
        query += ' ORDER BY local, nombre'

        cur.execute(query, params)
        rows = cur.fetchall()

        return jsonify({
            'total': len(rows),
            'data': rows
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/panel/borrar-stock', methods=['POST'])
def panel_borrar_stock():
    """Pone cantidad=NULL para fecha/bodega. NO toca conteos."""
    data = request.get_json()
    fecha = data.get('fecha')
    bodega = data.get('bodega', '')
    if not fecha:
        return jsonify({'error': 'Falta fecha'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Contar afectados
        q_count = """
            SELECT COUNT(*) as cnt FROM goti.inventario_ciego_conteos
            WHERE fecha = %s AND cantidad IS NOT NULL
        """
        params = [fecha]
        if bodega:
            q_count += ' AND local = %s'
            params.append(bodega)

        cur.execute(q_count, params)
        count = cur.fetchone()['cnt']

        if count == 0:
            return jsonify({'affected': 0, 'message': 'No hay registros con stock para esa fecha'})

        # Ejecutar UPDATE
        q_update = """
            UPDATE goti.inventario_ciego_conteos
            SET cantidad = NULL
            WHERE fecha = %s AND cantidad IS NOT NULL
        """
        params2 = [fecha]
        if bodega:
            q_update += ' AND local = %s'
            params2.append(bodega)

        cur.execute(q_update, params2)
        affected = cur.rowcount
        conn.commit()

        return jsonify({
            'affected': affected,
            'message': f'Stock borrado: {affected} registros actualizados'
        })
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/panel/contar-stock', methods=['GET'])
def panel_contar_stock():
    """Cuenta registros con stock para preview antes de borrar"""
    fecha = request.args.get('fecha')
    bodega = request.args.get('bodega', '')
    if not fecha:
        return jsonify({'error': 'Falta fecha'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT COUNT(*) as cnt FROM goti.inventario_ciego_conteos
            WHERE fecha = %s AND cantidad IS NOT NULL
        """
        params = [fecha]
        if bodega:
            query += ' AND local = %s'
            params.append(bodega)

        cur.execute(query, params)
        count = cur.fetchone()['cnt']

        return jsonify({'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


# ==================== ASIGNACION SEMANAL ====================

@app.route('/api/semanas', methods=['GET'])
def listar_semanas():
    """Lista semanas de inventario para una bodega"""
    local = request.args.get('local')
    if not local:
        return jsonify({'error': 'Falta parametro local'}), 400

    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT s.*,
                (SELECT COUNT(DISTINCT c.codigo)
                 FROM goti.inventario_ciego_conteos c
                 WHERE c.local = s.local
                   AND c.fecha BETWEEN s.fecha_inicio AND s.fecha_fin
                   AND (c.cantidad_contada IS NOT NULL OR c.cantidad_contada_2 IS NOT NULL)
                ) as total_productos,
                COALESCE((SELECT SUM(ap.monto)
                 FROM goti.asignacion_semanal a
                 JOIN goti.asignacion_semanal_personas ap ON ap.asignacion_semanal_id = a.id
                 WHERE a.semana_id = s.id
                ), 0) as total_asignado
            FROM goti.semanas_inventario s
            WHERE s.local = %s
        """
        params = [local]

        if fecha_desde:
            query += ' AND s.fecha_inicio >= %s'
            params.append(fecha_desde)
        if fecha_hasta:
            query += ' AND s.fecha_fin <= %s'
            params.append(fecha_hasta)

        query += ' ORDER BY s.fecha_inicio DESC'
        cur.execute(query, params)
        semanas = cur.fetchall()

        # Convert dates to strings
        for s in semanas:
            s['fecha_inicio'] = str(s['fecha_inicio'])
            s['fecha_fin'] = str(s['fecha_fin'])
            if s.get('cerrada_at'):
                s['cerrada_at'] = str(s['cerrada_at'])
            if s.get('created_at'):
                s['created_at'] = str(s['created_at'])

        return jsonify(semanas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/crear', methods=['POST'])
def crear_semana():
    """Crea o retorna una semana de inventario"""
    data = request.get_json()
    local = data.get('local')
    fecha_inicio = data.get('fecha_inicio')

    if not local or not fecha_inicio:
        return jsonify({'error': 'Faltan parametros local y fecha_inicio'}), 400

    from datetime import datetime, timedelta
    try:
        dt_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'fecha_inicio debe ser formato YYYY-MM-DD'}), 400

    # Validar que sea lunes (ISO weekday 1)
    if dt_inicio.isoweekday() != 1:
        return jsonify({'error': 'fecha_inicio debe ser un lunes'}), 400

    dt_fin = dt_inicio + timedelta(days=6)  # domingo

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Verificar si ya existe
        cur.execute("""
            SELECT * FROM goti.semanas_inventario
            WHERE fecha_inicio = %s AND local = %s
        """, (dt_inicio, local))
        existing = cur.fetchone()

        if existing:
            existing['fecha_inicio'] = str(existing['fecha_inicio'])
            existing['fecha_fin'] = str(existing['fecha_fin'])
            if existing.get('cerrada_at'):
                existing['cerrada_at'] = str(existing['cerrada_at'])
            if existing.get('created_at'):
                existing['created_at'] = str(existing['created_at'])
            return jsonify(existing)

        # Verificar que no haya otra semana abierta para este local
        cur.execute("""
            SELECT id, fecha_inicio, fecha_fin FROM goti.semanas_inventario
            WHERE local = %s AND estado = 'abierta'
        """, (local,))
        abierta = cur.fetchone()

        if abierta:
            return jsonify({
                'error': f'Ya existe una semana abierta para {local} ({abierta["fecha_inicio"]} - {abierta["fecha_fin"]}). Cierre primero antes de crear otra.'
            }), 409

        cur.execute("""
            INSERT INTO goti.semanas_inventario (fecha_inicio, fecha_fin, local)
            VALUES (%s, %s, %s)
            RETURNING *
        """, (dt_inicio, dt_fin, local))
        nueva = cur.fetchone()
        conn.commit()

        nueva['fecha_inicio'] = str(nueva['fecha_inicio'])
        nueva['fecha_fin'] = str(nueva['fecha_fin'])
        if nueva.get('created_at'):
            nueva['created_at'] = str(nueva['created_at'])

        return jsonify(nueva), 201
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/<int:semana_id>/diferencias', methods=['GET'])
def diferencias_semana(semana_id):
    """Obtiene diferencias semanales de productos para una semana"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Obtener datos de la semana
        cur.execute("""
            SELECT * FROM goti.semanas_inventario WHERE id = %s
        """, (semana_id,))
        semana = cur.fetchone()
        if not semana:
            return jsonify({'error': 'Semana no encontrada'}), 404

        fecha_inicio = semana['fecha_inicio']
        fecha_fin = semana['fecha_fin']
        local = semana['local']

        # Netear diferencias diarias por producto en la semana
        # Cada día: diferencia = conteo - sistema. Neto = suma de diferencias diarias.
        # Producto queda "justificado" si CUALQUIER día de la semana tiene justificado=TRUE
        cur.execute("""
            WITH diferencias_diarias AS (
                SELECT
                    codigo, nombre, unidad, fecha,
                    cantidad as stock_sistema,
                    COALESCE(cantidad_contada_2, cantidad_contada) as contado,
                    COALESCE(cantidad_contada_2, cantidad_contada) - cantidad as dif_dia,
                    COALESCE(costo_unitario, 0) as costo_unitario,
                    COALESCE(corregido, FALSE) as corregido,
                    COALESCE(justificado, FALSE) as justificado
                FROM goti.inventario_ciego_conteos
                WHERE local = %s AND fecha BETWEEN %s AND %s
                  AND COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
            )
            SELECT
                codigo,
                nombre,
                unidad,
                SUM(dif_dia) as diferencia,
                AVG(costo_unitario) as costo_unitario,
                COUNT(*) as dias_contados,
                BOOL_OR(justificado) as justificado,
                BOOL_OR(corregido) as tiene_correccion,
                json_agg(json_build_object(
                    'fecha', fecha,
                    'stock', stock_sistema,
                    'contado', contado,
                    'dif', dif_dia,
                    'corregido', corregido,
                    'justificado', justificado
                ) ORDER BY fecha) as detalle_diario
            FROM diferencias_diarias
            GROUP BY codigo, nombre, unidad
            HAVING SUM(dif_dia) != 0
            ORDER BY nombre
        """, (local, fecha_inicio, fecha_fin))
        diferencias = cur.fetchall()

        # Serializar datos — costo incluye 20% por costos indirectos (no visible al usuario)
        FACTOR_COSTO_INDIRECTO = 1.20
        for d in diferencias:
            d['diferencia'] = float(d['diferencia']) if d['diferencia'] else 0
            costo_base = float(d['costo_unitario']) if d['costo_unitario'] else 0
            d['costo_unitario'] = round(costo_base * FACTOR_COSTO_INDIRECTO, 4)
            if d.get('detalle_diario'):
                for dd in d['detalle_diario']:
                    dd['fecha'] = str(dd['fecha'])
                    dd['stock'] = float(dd['stock']) if dd['stock'] else 0
                    dd['contado'] = float(dd['contado']) if dd['contado'] else 0
                    dd['dif'] = float(dd['dif']) if dd['dif'] else 0

        # Obtener asignaciones existentes para esta semana
        cur.execute("""
            SELECT a.id, a.codigo, a.nombre, a.unidad, a.diferencia_semanal, a.costo_unitario,
                   json_agg(json_build_object(
                       'id', ap.id,
                       'persona', ap.persona,
                       'cantidad', ap.cantidad,
                       'monto', ap.monto
                   )) FILTER (WHERE ap.id IS NOT NULL) as personas
            FROM goti.asignacion_semanal a
            LEFT JOIN goti.asignacion_semanal_personas ap
                ON ap.asignacion_semanal_id = a.id
            WHERE a.semana_id = %s
            GROUP BY a.id, a.codigo, a.nombre, a.unidad, a.diferencia_semanal, a.costo_unitario
        """, (semana_id,))
        asignaciones = cur.fetchall()

        # Mapear asignaciones por codigo
        asig_map = {}
        for a in asignaciones:
            asig_map[a['codigo']] = {
                'id': a['id'],
                'diferencia_semanal': a['diferencia_semanal'],
                'costo_unitario': a['costo_unitario'],
                'personas': a['personas'] or []
            }

        # Combinar diferencias con asignaciones
        resultado = []
        for d in diferencias:
            item = dict(d)
            if d['codigo'] in asig_map:
                item['asignacion'] = asig_map[d['codigo']]
            else:
                item['asignacion'] = None
            resultado.append(item)

        semana_info = {
            'id': semana['id'],
            'fecha_inicio': str(semana['fecha_inicio']),
            'fecha_fin': str(semana['fecha_fin']),
            'local': semana['local'],
            'estado': semana['estado']
        }

        return jsonify({
            'semana': semana_info,
            'diferencias': resultado
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/<int:semana_id>/asignar', methods=['POST'])
def asignar_semana(semana_id):
    """Guarda asignaciones semanales de diferencias"""
    data = request.get_json()
    asignaciones = data.get('asignaciones', [])

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Verificar que la semana existe y esta abierta
        cur.execute("""
            SELECT * FROM goti.semanas_inventario WHERE id = %s
        """, (semana_id,))
        semana = cur.fetchone()
        if not semana:
            return jsonify({'error': 'Semana no encontrada'}), 404
        if semana['estado'] != 'abierta':
            return jsonify({'error': 'La semana esta cerrada, no se pueden modificar asignaciones'}), 400

        # Borrar asignaciones previas de esta semana
        cur.execute("""
            DELETE FROM goti.asignacion_semanal_personas
            WHERE asignacion_semanal_id IN (
                SELECT id FROM goti.asignacion_semanal WHERE semana_id = %s
            )
        """, (semana_id,))
        cur.execute("""
            DELETE FROM goti.asignacion_semanal WHERE semana_id = %s
        """, (semana_id,))

        # Insertar nuevas asignaciones
        total_insertadas = 0
        for asig in asignaciones:
            cur.execute("""
                INSERT INTO goti.asignacion_semanal
                    (semana_id, codigo, nombre, unidad, local, diferencia_semanal, costo_unitario)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                semana_id,
                asig.get('codigo'),
                asig.get('nombre'),
                asig.get('unidad'),
                semana['local'],
                asig.get('diferencia_semanal', 0),
                asig.get('costo_unitario', 0)
            ))
            asig_id = cur.fetchone()['id']

            for persona in asig.get('personas', []):
                cantidad = persona.get('cantidad', 0)
                costo = asig.get('costo_unitario', 0)  # ya viene con 20% desde frontend
                monto = float(cantidad) * float(costo) if cantidad and costo else 0
                cur.execute("""
                    INSERT INTO goti.asignacion_semanal_personas
                        (asignacion_semanal_id, persona, cantidad, monto)
                    VALUES (%s, %s, %s, %s)
                """, (asig_id, persona.get('persona'), cantidad, round(monto, 2)))

            total_insertadas += 1

        conn.commit()
        return jsonify({
            'ok': True,
            'message': f'{total_insertadas} asignaciones guardadas para semana {semana_id}'
        })
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/<int:semana_id>/cerrar', methods=['POST'])
def cerrar_semana(semana_id):
    """Cierra una semana de inventario"""
    data = request.get_json() or {}
    cerrada_por = data.get('cerrada_por', 'sistema')

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            SELECT estado FROM goti.semanas_inventario WHERE id = %s
        """, (semana_id,))
        semana = cur.fetchone()
        if not semana:
            return jsonify({'error': 'Semana no encontrada'}), 404
        if semana['estado'] != 'abierta':
            return jsonify({'error': 'La semana ya esta cerrada'}), 400

        cur.execute("""
            UPDATE goti.semanas_inventario
            SET estado = 'cerrada', cerrada_por = %s, cerrada_at = NOW()
            WHERE id = %s
            RETURNING *
        """, (cerrada_por, semana_id))
        updated = cur.fetchone()
        conn.commit()

        updated['fecha_inicio'] = str(updated['fecha_inicio'])
        updated['fecha_fin'] = str(updated['fecha_fin'])
        if updated.get('cerrada_at'):
            updated['cerrada_at'] = str(updated['cerrada_at'])
        if updated.get('created_at'):
            updated['created_at'] = str(updated['created_at'])

        return jsonify({'ok': True, 'semana': updated})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/<int:semana_id>', methods=['DELETE'])
def eliminar_semana(semana_id):
    """Elimina una semana (abierta o cerrada) y todas sus asignaciones (solo admin)"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("SELECT estado FROM goti.semanas_inventario WHERE id = %s", (semana_id,))
        semana = cur.fetchone()
        if not semana:
            return jsonify({'error': 'Semana no encontrada'}), 404

        # Admin puede eliminar tanto abiertas como cerradas
        # Al eliminar, los productos asignados quedan sin responsable y deben ser reasignados

        # Eliminar asignaciones de personas
        cur.execute("""
            DELETE FROM goti.asignacion_semanal_personas
            WHERE asignacion_semanal_id IN (
                SELECT id FROM goti.asignacion_semanal WHERE semana_id = %s
            )
        """, (semana_id,))
        # Eliminar asignaciones
        cur.execute("DELETE FROM goti.asignacion_semanal WHERE semana_id = %s", (semana_id,))
        # Eliminar semana
        cur.execute("DELETE FROM goti.semanas_inventario WHERE id = %s", (semana_id,))
        conn.commit()

        return jsonify({'success': True, 'estado_previo': semana['estado']})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/semanas/<int:semana_id>/reabrir', methods=['POST'])
def reabrir_semana(semana_id):
    """Reabre una semana cerrada (solo admin)"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            SELECT estado FROM goti.semanas_inventario WHERE id = %s
        """, (semana_id,))
        semana = cur.fetchone()
        if not semana:
            return jsonify({'error': 'Semana no encontrada'}), 404
        if semana['estado'] != 'cerrada':
            return jsonify({'error': 'La semana ya esta abierta'}), 400

        cur.execute("""
            UPDATE goti.semanas_inventario
            SET estado = 'abierta', cerrada_por = NULL, cerrada_at = NULL
            WHERE id = %s
            RETURNING *
        """, (semana_id,))
        updated = cur.fetchone()
        conn.commit()

        updated['fecha_inicio'] = str(updated['fecha_inicio'])
        updated['fecha_fin'] = str(updated['fecha_fin'])
        if updated.get('created_at'):
            updated['created_at'] = str(updated['created_at'])

        return jsonify({'ok': True, 'semana': updated})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/pendientes', methods=['GET'])
def semanas_pendientes():
    """Retorna semanas abiertas cuyo periodo ya termino (para recordatorios)"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            SELECT s.*
            FROM goti.semanas_inventario s
            WHERE s.estado = 'abierta'
              AND s.fecha_fin < CURRENT_DATE
            ORDER BY s.fecha_fin ASC
        """)
        semanas = cur.fetchall()

        for s in semanas:
            s['fecha_inicio'] = str(s['fecha_inicio'])
            s['fecha_fin'] = str(s['fecha_fin'])
            if s.get('created_at'):
                s['created_at'] = str(s['created_at'])

        return jsonify(semanas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/resumen-persona', methods=['GET'])
def resumen_persona_semanal():
    """Resumen de asignaciones por persona a traves de semanas cerradas"""
    local = request.args.get('local')
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    if not local:
        return jsonify({'error': 'Falta parametro local'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT ap.persona,
                   SUM(ap.cantidad) as total_cantidad,
                   SUM(ap.monto) as total_monto,
                   COUNT(DISTINCT a.semana_id) as semanas_count
            FROM goti.asignacion_semanal_personas ap
            JOIN goti.asignacion_semanal a ON a.id = ap.asignacion_semanal_id
            JOIN goti.semanas_inventario s ON s.id = a.semana_id
            WHERE s.local = %s AND s.estado = 'cerrada'
        """
        params = [local]

        if fecha_desde:
            query += ' AND s.fecha_inicio >= %s'
            params.append(fecha_desde)
        if fecha_hasta:
            query += ' AND s.fecha_fin <= %s'
            params.append(fecha_hasta)

        query += ' GROUP BY ap.persona ORDER BY total_monto DESC'
        cur.execute(query, params)
        resumen = cur.fetchall()

        return jsonify(resumen)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


# ============================================================
# MODULO: Cruce Operativo "CUADRAR" - Boton + Worker local
# ============================================================
# Flujo: Panel web -> POST /solicitar -> tarea pendiente
#        Worker PC FINANZAS -> GET /pendientes (cada 15s) -> toma tarea
#        Worker descarga Contifico, calcula cruce -> POST /resultado
#        Panel web -> GET /estado/<id> (polling) -> muestra resultado

# Token simple para autenticar al worker (env var)
WORKER_TOKEN = os.environ.get('CRUCE_WORKER_TOKEN', 'worker-foodix-2026-7K3xR9pL2qN8mZ4w')


@app.route('/api/cruce-op/solicitar', methods=['POST'])
def cruce_op_solicitar():
    """Llamado desde el panel cuando el usuario presiona CUADRAR.
    Crea una tarea pendiente que el worker tomara.
    Si ya existe completado, solo admin puede re-ejecutar.
    Si ya hay una pendiente o en proceso, devuelve esa misma."""
    data = request.json or {}
    bodega = data.get('bodega')
    fecha_toma = data.get('fecha_toma')
    fecha_corte = data.get('fecha_corte_contifico') or fecha_toma  # por defecto = fecha_toma
    usuario = data.get('usuario', 'panel')
    rol = data.get('rol', '')

    if bodega not in ('bodega_principal', 'materia_prima', 'planta'):
        return jsonify({'error': 'bodega invalida'}), 400
    if not fecha_toma:
        return jsonify({'error': 'fecha_toma requerida'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Ver si ya existe alguna ejecucion para esta bodega+fecha
        cur.execute("""
            SELECT id, estado FROM goti.cruce_operativo_ejecuciones
            WHERE bodega = %s AND fecha_toma = %s
            ORDER BY COALESCE(solicitado_at, timestamp_deteccion) DESC LIMIT 1
        """, (bodega, fecha_toma))
        existente = cur.fetchone()

        if existente and existente['estado'] in ('pendiente', 'en_proceso'):
            # Ya se esta procesando, devolver la misma
            return jsonify({'id': existente['id'], 'estado': existente['estado'], 'reused': True})

        if existente and existente['estado'] == 'completado':
            # Solo admin puede re-ejecutar un cruce ya completado
            if rol != 'admin':
                return jsonify({'error': 'Este cruce ya fue ejecutado. Solo el administrador puede re-ejecutarlo.', 'ya_completado': True}), 409
            # Admin: resetear
            cur.execute("DELETE FROM goti.cruce_operativo_detalle WHERE ejecucion_id = %s", (existente['id'],))
            cur.execute("""
                UPDATE goti.cruce_operativo_ejecuciones
                SET estado='pendiente', solicitado_por=%s, solicitado_at=NOW(),
                    fecha_corte_contifico=%s,
                    worker_lock=NULL, error_msg=NULL,
                    timestamp_descarga=NULL, timestamp_cruce=NULL,
                    total_productos_toma=NULL, total_productos_contifico=NULL,
                    total_cruzados=NULL, total_con_diferencia=NULL, valor_total_dif=NULL
                WHERE id = %s
            """, (usuario, fecha_corte, existente['id']))
            conn.commit()
            return jsonify({'id': existente['id'], 'estado': 'pendiente', 'reset': True})

        if existente:
            # Estado error: cualquiera puede reintentar
            cur.execute("DELETE FROM goti.cruce_operativo_detalle WHERE ejecucion_id = %s", (existente['id'],))
            cur.execute("""
                UPDATE goti.cruce_operativo_ejecuciones
                SET estado='pendiente', solicitado_por=%s, solicitado_at=NOW(),
                    fecha_corte_contifico=%s,
                    worker_lock=NULL, error_msg=NULL,
                    timestamp_descarga=NULL, timestamp_cruce=NULL,
                    total_productos_toma=NULL, total_productos_contifico=NULL,
                    total_cruzados=NULL, total_con_diferencia=NULL, valor_total_dif=NULL
                WHERE id = %s
            """, (usuario, fecha_corte, existente['id']))
            conn.commit()
            return jsonify({'id': existente['id'], 'estado': 'pendiente', 'reset': True})

        # No existe: crear nueva
        cur.execute("""
            INSERT INTO goti.cruce_operativo_ejecuciones
            (bodega, fecha_toma, fecha_corte_contifico, estado, solicitado_por, solicitado_at)
            VALUES (%s, %s, %s, 'pendiente', %s, NOW())
            RETURNING id
        """, (bodega, fecha_toma, fecha_corte, usuario))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'id': new_id, 'estado': 'pendiente'})
    except Exception as e:
        print(f"Error en /api/cruce-op/solicitar: {e}")
        if conn: conn.rollback()
        return jsonify({'error': 'Error interno del servidor', 'detalle': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/eliminar/<int:ejec_id>', methods=['DELETE'])
def cruce_op_eliminar(ejec_id):
    """Elimina una ejecucion y su detalle. Llamado desde el panel."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.cruce_operativo_detalle WHERE ejecucion_id = %s", (ejec_id,))
        cur.execute("DELETE FROM goti.cruce_operativo_ejecuciones WHERE id = %s", (ejec_id,))
        conn.commit()
        return jsonify({'ok': True, 'eliminados': cur.rowcount})
    except Exception as e:
        print(f"Error en /api/cruce-op/eliminar: {e}")
        if conn: conn.rollback()
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/pendientes', methods=['GET'])
def cruce_op_pendientes():
    """Llamado por el worker. Devuelve tareas pendientes y las marca como en_proceso."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401

    worker_id = request.args.get('worker_id', 'pc-finanzas')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Marca atomicamente las pendientes como en_proceso para este worker
        cur.execute("""
            UPDATE goti.cruce_operativo_ejecuciones
            SET estado = 'en_proceso',
                worker_lock = %s,
                timestamp_descarga = NOW()
            WHERE id IN (
                SELECT id FROM goti.cruce_operativo_ejecuciones
                WHERE estado = 'pendiente'
                ORDER BY solicitado_at ASC
                LIMIT 5
                FOR UPDATE SKIP LOCKED
            )
            RETURNING id, bodega, fecha_toma, fecha_corte_contifico, solicitado_por, solicitado_at
        """, (worker_id,))
        rows = cur.fetchall()
        conn.commit()
        result = [{
            'id': r['id'],
            'bodega': r['bodega'],
            'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
            'fecha_corte_contifico': r['fecha_corte_contifico'].isoformat() if r['fecha_corte_contifico'] else (r['fecha_toma'].isoformat() if r['fecha_toma'] else None),
            'solicitado_por': r['solicitado_por'],
        } for r in rows]
        return jsonify(result)
    except Exception as e:
        print(f"Error en /api/cruce-op/pendientes: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/resultado', methods=['POST'])
def cruce_op_resultado():
    """Llamado por el worker al terminar. Inserta detalle y marca completado/error."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401

    data = request.json or {}
    ejec_id = data.get('id')
    estado = data.get('estado', 'completado')  # 'completado' o 'error'
    error_msg = data.get('error_msg')
    detalle = data.get('detalle', [])
    resumen = data.get('resumen', {})

    if not ejec_id:
        return jsonify({'error': 'id requerido'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        if estado == 'error':
            cur.execute("""
                UPDATE goti.cruce_operativo_ejecuciones
                SET estado = 'error', error_msg = %s, timestamp_cruce = NOW()
                WHERE id = %s
            """, (error_msg, ejec_id))
            conn.commit()
            return jsonify({'ok': True})

        # Borrar detalle previo si existiera
        cur.execute("DELETE FROM goti.cruce_operativo_detalle WHERE ejecucion_id = %s", (ejec_id,))

        # Insertar detalle
        if detalle:
            for d in detalle:
                cur.execute("""
                    INSERT INTO goti.cruce_operativo_detalle
                    (ejecucion_id, codigo, nombre, categoria, unidad, unidad_toma, factor,
                     unidad_destino, cantidad_toma, cantidad_sistema, diferencia,
                     costo_unitario, valor_diferencia, tipo_abc, origen)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    ejec_id, d.get('codigo'), d.get('nombre'), d.get('categoria'),
                    d.get('unidad_destino'), d.get('unidad_toma'), d.get('factor'),
                    d.get('unidad_destino'), d.get('cantidad_toma'), d.get('cantidad_sistema'),
                    d.get('diferencia'), d.get('costo_unitario'), d.get('valor_diferencia'),
                    d.get('tipo_abc'), d.get('origen', 'cruce_operativo')
                ))

        # Update ejecucion
        cur.execute("""
            UPDATE goti.cruce_operativo_ejecuciones
            SET estado = 'completado',
                total_productos_toma = %s,
                total_productos_contifico = %s,
                total_cruzados = %s,
                total_con_diferencia = %s,
                valor_total_dif = %s,
                timestamp_cruce = NOW()
            WHERE id = %s
        """, (
            resumen.get('total_productos_toma'),
            resumen.get('total_productos_contifico'),
            resumen.get('total_cruzados'),
            resumen.get('total_con_diferencia'),
            resumen.get('valor_total_dif'),
            ejec_id
        ))
        conn.commit()
        return jsonify({'ok': True, 'detalles_insertados': len(detalle)})
    except Exception as e:
        print(f"Error en /api/cruce-op/resultado: {e}")
        if conn:
            conn.rollback()
        return jsonify({'error': 'Error interno del servidor', 'detalle': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/estado/<int:ejec_id>', methods=['GET'])
def cruce_op_estado(ejec_id):
    """Polling desde el panel para saber estado de una ejecucion."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, bodega, fecha_toma, estado, solicitado_por, solicitado_at,
                   timestamp_descarga, timestamp_cruce, error_msg,
                   total_productos_toma, total_productos_contifico, total_cruzados,
                   total_con_diferencia, valor_total_dif
            FROM goti.cruce_operativo_ejecuciones WHERE id = %s
        """, (ejec_id,))
        r = cur.fetchone()
        if not r:
            return jsonify({'error': 'no encontrado'}), 404
        return jsonify({
            'id': r['id'],
            'bodega': r['bodega'],
            'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
            'estado': r['estado'],
            'solicitado_por': r['solicitado_por'],
            'solicitado_at': r['solicitado_at'].isoformat() if r['solicitado_at'] else None,
            'timestamp_descarga': r['timestamp_descarga'].isoformat() if r['timestamp_descarga'] else None,
            'timestamp_cruce': r['timestamp_cruce'].isoformat() if r['timestamp_cruce'] else None,
            'error_msg': r['error_msg'],
            'total_productos_toma': r['total_productos_toma'],
            'total_productos_contifico': r['total_productos_contifico'],
            'total_cruzados': r['total_cruzados'],
            'total_con_diferencia': r['total_con_diferencia'],
            'valor_total_dif': float(r['valor_total_dif']) if r['valor_total_dif'] is not None else None,
        })
    except Exception as e:
        print(f"Error en /api/cruce-op/estado: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/fechas-disponibles', methods=['GET'])
def cruce_op_fechas():
    """Devuelve las fechas con toma fisica disponibles para una bodega."""
    bodega = request.args.get('bodega')
    tablas = {
        'bodega_principal': 'public.toma_bodega',
        'materia_prima':    'public.toma_materiaprima',
        'planta':           'public.toma_planta',
    }
    if bodega not in tablas:
        return jsonify({'error': 'bodega invalida'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(f"""
            SELECT fecha, COUNT(*) AS productos
            FROM {tablas[bodega]}
            WHERE fecha IS NOT NULL
            GROUP BY fecha ORDER BY fecha DESC
        """)
        rows = cur.fetchall()
        return jsonify([{
            'fecha': r['fecha'].isoformat(),
            'productos': r['productos']
        } for r in rows])
    except Exception as e:
        print(f"Error en /api/cruce-op/fechas-disponibles: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


# ============================================================
# MODULO: Carga Toma Fisica a Contifico
# ============================================================
# Flujo: Panel web -> POST /solicitar -> tarea pendiente
#        Worker PC FINANZAS -> GET /pendientes-carga (cada 15s) -> toma tarea
#        Worker abre Contifico, llena formulario toma fisica -> POST /resultado-carga
#        Panel web -> GET /estado-carga/<id> (polling) -> muestra resultado
#        UNIQUE(bodega, fecha_toma) -> no permite cargar dos veces

@app.route('/api/carga-contifico/fechas-con-cruce', methods=['GET'])
def carga_contifico_fechas_con_cruce():
    """Devuelve las fechas que tienen cruce operativo completado para una bodega."""
    bodega = request.args.get('bodega')
    if bodega not in ('bodega_principal', 'materia_prima', 'planta'):
        return jsonify({'error': 'bodega invalida'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT fecha_toma, total_cruzados, total_con_diferencia, valor_total_dif
            FROM goti.cruce_operativo_ejecuciones
            WHERE bodega = %s AND estado = 'completado'
            ORDER BY fecha_toma DESC
        """, (bodega,))
        rows = cur.fetchall()
        return jsonify([{
            'fecha': r['fecha_toma'].isoformat(),
            'cruzados': r['total_cruzados'],
            'con_dif': r['total_con_diferencia'],
            'valor_dif': float(r['valor_total_dif']) if r['valor_total_dif'] else 0,
        } for r in rows])
    except Exception as e:
        print(f"Error en /api/carga-contifico/fechas-con-cruce: {e}")
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/carga-contifico/verificar', methods=['GET'])
def carga_contifico_verificar():
    """Verifica si ya existe una carga completada para bodega+fecha."""
    bodega = request.args.get('bodega')
    fecha = request.args.get('fecha')
    if not bodega or not fecha:
        return jsonify({'error': 'bodega y fecha requeridos'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, estado, solicitado_at, timestamp_fin, total_productos, productos_ok, productos_error
            FROM goti.carga_contifico_ejecuciones
            WHERE bodega = %s AND fecha_toma = %s
        """, (bodega, fecha))
        row = cur.fetchone()
        if not row:
            return jsonify({'existe': False, 'cargado': False})
        return jsonify({
            'existe': True,
            'cargado': row['estado'] == 'completado',
            'estado': row['estado'],
            'id': row['id'],
            'solicitado_at': row['solicitado_at'].isoformat() if row['solicitado_at'] else None,
            'timestamp_fin': row['timestamp_fin'].isoformat() if row['timestamp_fin'] else None,
            'total_productos': row['total_productos'],
            'productos_ok': row['productos_ok'],
            'productos_error': row['productos_error'],
        })
    except Exception as e:
        print(f"Error en /api/carga-contifico/verificar: {e}")
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/carga-contifico/solicitar', methods=['POST'])
def carga_contifico_solicitar():
    """Crea tarea de carga. Si ya esta completada, solo admin puede re-ejecutar."""
    data = request.json or {}
    bodega = data.get('bodega')
    fecha_toma = data.get('fecha_toma')
    usuario = data.get('usuario', 'panel')
    rol = data.get('rol', '')

    if bodega not in ('bodega_principal', 'materia_prima', 'planta'):
        return jsonify({'error': 'bodega invalida'}), 400
    if not fecha_toma:
        return jsonify({'error': 'fecha_toma requerida'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, estado FROM goti.carga_contifico_ejecuciones
            WHERE bodega = %s AND fecha_toma = %s
        """, (bodega, fecha_toma))
        existente = cur.fetchone()

        if existente:
            if existente['estado'] == 'completado':
                if rol != 'admin':
                    return jsonify({'error': 'Ya fue cargado a Contifico. Solo el administrador puede re-ejecutar.', 'ya_cargado': True}), 409
                # Admin: resetear para re-ejecutar
                cur.execute("""
                    UPDATE goti.carga_contifico_ejecuciones
                    SET estado='pendiente', solicitado_por=%s, solicitado_at=NOW(),
                        worker_lock=NULL, error_msg=NULL, timestamp_inicio=NULL, timestamp_fin=NULL,
                        total_productos=NULL, productos_ok=NULL, productos_error=NULL, productos_error_lista=NULL
                    WHERE id = %s
                """, (usuario, existente['id']))
                conn.commit()
                return jsonify({'id': existente['id'], 'estado': 'pendiente', 'reset': True})
            if existente['estado'] in ('pendiente', 'en_proceso'):
                return jsonify({'id': existente['id'], 'estado': existente['estado'], 'reused': True})
            # Estado error: resetear para reintentar
            cur.execute("""
                UPDATE goti.carga_contifico_ejecuciones
                SET estado='pendiente', solicitado_por=%s, solicitado_at=NOW(),
                    worker_lock=NULL, error_msg=NULL, timestamp_inicio=NULL, timestamp_fin=NULL,
                    total_productos=NULL, productos_ok=NULL, productos_error=NULL, productos_error_lista=NULL
                WHERE id = %s
            """, (usuario, existente['id']))
            conn.commit()
            return jsonify({'id': existente['id'], 'estado': 'pendiente', 'reset': True})

        # No existe: crear nueva
        cur.execute("""
            INSERT INTO goti.carga_contifico_ejecuciones
            (bodega, fecha_toma, estado, solicitado_por, solicitado_at)
            VALUES (%s, %s, 'pendiente', %s, NOW())
            RETURNING id
        """, (bodega, fecha_toma, usuario))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'id': new_id, 'estado': 'pendiente'})
    except Exception as e:
        print(f"Error en /api/carga-contifico/solicitar: {e}")
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/carga-contifico/pendientes', methods=['GET'])
def carga_contifico_pendientes():
    """Llamado por el worker. Devuelve tareas pendientes de carga y las marca en_proceso."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401

    worker_id = request.args.get('worker_id', 'pc-finanzas')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.carga_contifico_ejecuciones
            SET estado = 'en_proceso', worker_lock = %s, timestamp_inicio = NOW()
            WHERE id IN (
                SELECT id FROM goti.carga_contifico_ejecuciones
                WHERE estado = 'pendiente'
                ORDER BY solicitado_at ASC
                LIMIT 1
                FOR UPDATE SKIP LOCKED
            )
            RETURNING id, bodega, fecha_toma, solicitado_por
        """, (worker_id,))
        rows = cur.fetchall()
        conn.commit()
        return jsonify([{
            'id': r['id'],
            'bodega': r['bodega'],
            'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
            'tipo': 'carga_contifico',
        } for r in rows])
    except Exception as e:
        print(f"Error en /api/carga-contifico/pendientes: {e}")
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/carga-contifico/resultado', methods=['POST'])
def carga_contifico_resultado():
    """Llamado por el worker al terminar la carga."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401

    data = request.json or {}
    ejec_id = data.get('id')
    estado = data.get('estado', 'completado')
    error_msg = data.get('error_msg')

    if not ejec_id:
        return jsonify({'error': 'id requerido'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.carga_contifico_ejecuciones
            SET estado = %s, timestamp_fin = NOW(),
                total_productos = %s, productos_ok = %s,
                productos_error = %s, productos_error_lista = %s,
                error_msg = %s
            WHERE id = %s
        """, (
            estado,
            data.get('total_productos'),
            data.get('productos_ok'),
            data.get('productos_error'),
            data.get('productos_error_lista'),
            error_msg,
            ejec_id
        ))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        print(f"Error en /api/carga-contifico/resultado: {e}")
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/carga-contifico/estado/<int:ejec_id>', methods=['GET'])
def carga_contifico_estado(ejec_id):
    """Polling desde el panel."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, bodega, fecha_toma, estado, solicitado_at,
                   timestamp_inicio, timestamp_fin, error_msg,
                   total_productos, productos_ok, productos_error, productos_error_lista
            FROM goti.carga_contifico_ejecuciones WHERE id = %s
        """, (ejec_id,))
        r = cur.fetchone()
        if not r:
            return jsonify({'error': 'no encontrado'}), 404
        return jsonify({
            'id': r['id'],
            'bodega': r['bodega'],
            'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
            'estado': r['estado'],
            'solicitado_at': r['solicitado_at'].isoformat() if r['solicitado_at'] else None,
            'timestamp_inicio': r['timestamp_inicio'].isoformat() if r['timestamp_inicio'] else None,
            'timestamp_fin': r['timestamp_fin'].isoformat() if r['timestamp_fin'] else None,
            'error_msg': r['error_msg'],
            'total_productos': r['total_productos'],
            'productos_ok': r['productos_ok'],
            'productos_error': r['productos_error'],
            'productos_error_lista': r['productos_error_lista'],
        })
    except Exception as e:
        print(f"Error en /api/carga-contifico/estado: {e}")
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


# ============================================================
# MODULO: Evaluacion Semanal por Local
# ============================================================

EVAL_LOCALES = [
    {'id': 'real_audiencia', 'nombre': 'Chios Real Audiencia'},
    {'id': 'floreana', 'nombre': 'Chios Floreana'},
    {'id': 'portugal', 'nombre': 'Chios Portugal'},
    {'id': 'santo_cachon_real', 'nombre': 'Santo Cachon Real'},
    {'id': 'santo_cachon_portugal', 'nombre': 'Santo Cachon Portugal'},
    {'id': 'simon_bolon', 'nombre': 'Simon Bolon'},
]

@app.route('/api/eval/categorias', methods=['GET'])
def eval_categorias():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id, nombre, descripcion, orden, criterios FROM goti.eval_categorias WHERE activa = TRUE ORDER BY orden")
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/eval/locales', methods=['GET'])
def eval_locales():
    return jsonify(EVAL_LOCALES)


@app.route('/api/eval/guardar', methods=['POST'])
def eval_guardar():
    """Guarda evaluacion semanal. Body: {local, semana_inicio, semana_fin, evaluaciones: [{categoria_id, puntaje, comentario}], evaluado_por}"""
    data = request.json or {}
    local = data.get('local')
    semana_inicio = data.get('semana_inicio')
    semana_fin = data.get('semana_fin')
    evaluaciones = data.get('evaluaciones', [])
    evaluado_por = data.get('evaluado_por', 'admin')

    if not local or not semana_inicio or not evaluaciones:
        return jsonify({'error': 'local, semana_inicio y evaluaciones requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        guardados = 0
        for ev in evaluaciones:
            cur.execute("""
                INSERT INTO goti.eval_semanal
                (local, semana_inicio, semana_fin, categoria_id, puntaje, comentario, evaluado_por, evaluado_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT (local, semana_inicio, categoria_id)
                DO UPDATE SET puntaje = EXCLUDED.puntaje, comentario = EXCLUDED.comentario,
                              evaluado_por = EXCLUDED.evaluado_por, evaluado_at = NOW()
            """, (local, semana_inicio, semana_fin, ev['categoria_id'], ev['puntaje'], ev.get('comentario', ''), evaluado_por))
            guardados += 1
        conn.commit()
        return jsonify({'ok': True, 'guardados': guardados})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/eval/semana', methods=['GET'])
def eval_semana():
    """Obtiene evaluaciones de una semana. Params: semana_inicio, local (opcional)"""
    semana = request.args.get('semana_inicio')
    local = request.args.get('local')
    if not semana:
        return jsonify({'error': 'semana_inicio requerido'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        if local:
            cur.execute("""
                SELECT e.id, e.local, e.categoria_id, c.nombre as categoria, e.puntaje, e.comentario, e.evaluado_por, e.evaluado_at
                FROM goti.eval_semanal e
                JOIN goti.eval_categorias c ON c.id = e.categoria_id
                WHERE e.semana_inicio = %s AND e.local = %s
                ORDER BY c.orden
            """, (semana, local))
        else:
            cur.execute("""
                SELECT e.id, e.local, e.categoria_id, c.nombre as categoria, e.puntaje, e.comentario, e.evaluado_por, e.evaluado_at
                FROM goti.eval_semanal e
                JOIN goti.eval_categorias c ON c.id = e.categoria_id
                WHERE e.semana_inicio = %s
                ORDER BY e.local, c.orden
            """, (semana,))
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/eval/ranking', methods=['GET'])
def eval_ranking():
    """Ranking de locales por promedio. Params: semana_inicio (opcional, default ultima disponible), ultimas_n (semanas a promediar, default 1)"""
    semana = request.args.get('semana_inicio')
    ultimas_n = int(request.args.get('ultimas_n', '1'))
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        if semana:
            cur.execute("""
                SELECT local, ROUND(AVG(puntaje)::numeric, 2) as promedio, COUNT(DISTINCT categoria_id) as categorias_evaluadas
                FROM goti.eval_semanal
                WHERE semana_inicio = %s
                GROUP BY local ORDER BY promedio DESC
            """, (semana,))
        else:
            cur.execute("""
                SELECT local, ROUND(AVG(puntaje)::numeric, 2) as promedio, COUNT(DISTINCT semana_inicio) as semanas
                FROM goti.eval_semanal
                WHERE semana_inicio >= (
                    SELECT MAX(semana_inicio) - interval '%s weeks' FROM goti.eval_semanal
                )
                GROUP BY local ORDER BY promedio DESC
            """ % max(1, ultimas_n))
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/eval/tendencia', methods=['GET'])
def eval_tendencia():
    """Tendencia historica por local. Params: local (opcional), limite (semanas, default 12)"""
    local = request.args.get('local')
    limite = int(request.args.get('limite', '12'))
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        if local:
            cur.execute("""
                SELECT semana_inicio, ROUND(AVG(puntaje)::numeric, 2) as promedio
                FROM goti.eval_semanal
                WHERE local = %s
                GROUP BY semana_inicio ORDER BY semana_inicio DESC LIMIT %s
            """, (local, limite))
        else:
            cur.execute("""
                SELECT local, semana_inicio, ROUND(AVG(puntaje)::numeric, 2) as promedio
                FROM goti.eval_semanal
                GROUP BY local, semana_inicio ORDER BY semana_inicio DESC, promedio DESC
                LIMIT %s
            """, (limite * 6,))
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/eval/semanas-disponibles', methods=['GET'])
def eval_semanas_disponibles():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT semana_inicio, semana_fin
            FROM goti.eval_semanal
            ORDER BY semana_inicio DESC LIMIT 52
        """)
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/evaluacion')
def evaluacion_page():
    return render_template('evaluacion.html')


# ============================================================
# MODULO: DEPOSITOS (lee desde AirTable)
# ============================================================
AIRTABLE_DEPOSITOS_TOKEN = os.environ.get('AIRTABLE_DEPOSITOS_TOKEN', '') or os.environ.get('AIRTABLE_TOKEN', '') or _b64.b64decode('cGF0d1owSHBiRlQ5RkNoNWQuZDQwY2ExZTZlNGViYWRlZWE5ZjJmZGYyZTAwM2FhOGMxMGIyMjAzYzkxZjg2OTk1YmRiOTgyMjYwOTkzMzM3YQ==').decode()
print(f"[INIT] AIRTABLE_DEPOSITOS_TOKEN: {'SET (' + str(len(AIRTABLE_DEPOSITOS_TOKEN)) + ' chars)' if AIRTABLE_DEPOSITOS_TOKEN else 'EMPTY'}")
AIRTABLE_DEPOSITOS_BASE = 'apppZXgUChlBLbVpR'
AIRTABLE_DEPOSITOS_TABLE = 'tbldo5QTH6bBpgYbx'
AIRTABLE_TIENDAS_TABLE = 'tblxloBdnbdsGcuKR'

_tiendas_cache = {}
_tiendas_cache_ts = 0
_responsables_cache = {}
_responsables_cache_ts = 0

AIRTABLE_RESPONSABLES_TABLE = 'tblR8NOjmbp70LmTK'

def _at_headers():
    return {'Authorization': f'Bearer {AIRTABLE_DEPOSITOS_TOKEN}'}

def _cargar_tiendas():
    global _tiendas_cache, _tiendas_cache_ts
    import time as _t
    if _tiendas_cache and (_t.time() - _tiendas_cache_ts) < 600:
        return _tiendas_cache
    try:
        import requests as req
        all_recs = []
        offset = None
        while True:
            params = {'pageSize': 100, 'fields[]': ['Código', 'Marca', 'Ubicación']}
            if offset: params['offset'] = offset
            r = req.get(f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_TIENDAS_TABLE}',
                headers=_at_headers(), params=params, timeout=15)
            if r.status_code != 200: break
            data = r.json()
            all_recs.extend(data.get('records', []))
            offset = data.get('offset')
            if not offset: break
        for rec in all_recs:
            _tiendas_cache[rec['id']] = {
                'codigo': rec['fields'].get('Código', rec['id']),
                'ubicacion': rec['fields'].get('Ubicación', ''),
                'marca': rec['fields'].get('Marca', ''),
            }
        _tiendas_cache_ts = _t.time()
    except Exception as e:
        print(f'Error cargando tiendas: {e}')
    return _tiendas_cache

def _cargar_responsables():
    global _responsables_cache, _responsables_cache_ts
    import time as _t
    if _responsables_cache and (_t.time() - _responsables_cache_ts) < 600:
        return _responsables_cache
    try:
        import requests as req
        all_recs = []
        offset = None
        while True:
            params = {'pageSize': 100, 'fields[]': ['Nombre', 'Fecha Ingreso', 'Cargo']}
            if offset: params['offset'] = offset
            r = req.get(f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_RESPONSABLES_TABLE}',
                headers=_at_headers(), params=params, timeout=15)
            if r.status_code != 200: break
            data = r.json()
            all_recs.extend(data.get('records', []))
            offset = data.get('offset')
            if not offset: break
        for rec in all_recs:
            _responsables_cache[rec['id']] = {
                'nombre': rec['fields'].get('Nombre', ''),
                'fecha_ingreso': rec['fields'].get('Fecha Ingreso', ''),
                'cargo': rec['fields'].get('Cargo', ''),
            }
        _responsables_cache_ts = _t.time()
    except Exception as e:
        print(f'Error cargando responsables: {e}')
    return _responsables_cache

def _resolver_local(local_ids):
    if not local_ids:
        return 'Sin local'
    tiendas = _cargar_tiendas()
    nombres = [tiendas.get(lid, {}).get('codigo', lid) if isinstance(tiendas.get(lid), dict) else tiendas.get(lid, lid) for lid in local_ids]
    return ', '.join(nombres)

def _resolver_local_detalle(local_ids):
    if not local_ids:
        return {'codigo': 'Sin local', 'ubicacion': '', 'marca': ''}
    tiendas = _cargar_tiendas()
    t = tiendas.get(local_ids[0], {})
    if isinstance(t, str):
        return {'codigo': t, 'ubicacion': '', 'marca': ''}
    return t

def _resolver_responsable(resp_ids):
    if not resp_ids:
        return {'nombre': 'Sin responsable', 'fecha_ingreso': '', 'cargo': ''}
    responsables = _cargar_responsables()
    return responsables.get(resp_ids[0], {'nombre': resp_ids[0], 'fecha_ingreso': '', 'cargo': ''})


# Cache de depositos: {clave: {data, timestamp}}
_depositos_cache = {}
_DEPOSITOS_CACHE_TTL = 300  # 5 minutos

def _depositos_cache_key(fecha_desde, fecha_hasta, estado, cuadre):
    return f'{fecha_desde}|{fecha_hasta}|{estado}|{cuadre}'

def _invalidar_cache_depositos():
    """Limpia todo el cache de depositos (llamar al hacer una accion)."""
    global _depositos_cache
    _depositos_cache = {}


@app.route('/api/depositos/listar', methods=['GET'])
def depositos_listar():
    """Lista depositos desde AirTable con filtros. Cache de 5 minutos."""
    import requests as req
    import time as _t
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    estado = request.args.get('estado', '')
    cuadre = request.args.get('cuadre', '')

    # Verificar cache
    cache_key = _depositos_cache_key(fecha_desde, fecha_hasta, estado, cuadre)
    cached = _depositos_cache.get(cache_key)
    if cached and (_t.time() - cached['ts']) < _DEPOSITOS_CACHE_TTL:
        return jsonify(cached['data'])

    try:
        # Construir formula de filtro
        filtros = []
        if fecha_desde:
            filtros.append(f"IS_ON_OR_AFTER({{Fecha}}, '{fecha_desde}')")
        if fecha_hasta:
            filtros.append(f"IS_ON_OR_BEFORE({{Fecha}}, '{fecha_hasta}')")
        if estado:
            filtros.append(f"{{Estado}} = '{estado}'")
        if cuadre:
            filtros.append(f"{{Estado De Cuadre}} = '{cuadre}'")

        params = {
            'pageSize': 100,
            'sort[0][field]': 'Fecha',
            'sort[0][direction]': 'desc',
            'fields[]': ['Fecha', 'Local', 'Responsable De Caja', 'Monto Contado',
                         'Monto A Recibir', 'Diferencia Contado Vs. Recibido',
                         'Secuencia De Caja', 'Número De Depósitos', 'Estado',
                         'Estado De Cuadre', 'Observación', 'Evidencia', 'Evidencia Del Déposito',
                         'Fecha Creación', 'Correo (from Responsable De Caja)'],
        }
        if filtros:
            params['filterByFormula'] = 'AND(' + ','.join(filtros) + ')'

        all_records = []
        offset = None
        while True:
            if offset:
                params['offset'] = offset
            r = req.get(f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_DEPOSITOS_TABLE}',
                headers=_at_headers(), params=params, timeout=20)
            if r.status_code != 200:
                return jsonify({'error': f'AirTable error: {r.status_code}'}), 500
            data = r.json()
            all_records.extend(data.get('records', []))
            offset = data.get('offset')
            if not offset or len(all_records) >= 500:
                break

        # Pre-cargar caches
        _cargar_tiendas()
        _cargar_responsables()

        # Resolver locales
        resultado = []
        for rec in all_records:
            f = rec['fields']
            evidencias = []
            for att in (f.get('Evidencia', [])):
                if isinstance(att, dict):
                    thumb = att.get('thumbnails', {}).get('large', {}).get('url', '')
                    evidencias.append({'url': att.get('url', ''), 'thumb': thumb, 'filename': att.get('filename', '')})
            evidencias_deposito = []
            for att in (f.get('Evidencia Del Déposito', [])):
                if isinstance(att, dict):
                    thumb = att.get('thumbnails', {}).get('large', {}).get('url', '')
                    evidencias_deposito.append({'url': att.get('url', ''), 'thumb': thumb, 'filename': att.get('filename', '')})

            local_det = _resolver_local_detalle(f.get('Local', []))
            resp_det = _resolver_responsable(f.get('Responsable De Caja', []))

            resultado.append({
                'id': rec['id'],
                'fecha': f.get('Fecha'),
                'local': local_det.get('codigo', ''),
                'local_ubicacion': local_det.get('ubicacion', ''),
                'local_marca': local_det.get('marca', ''),
                'responsable': resp_det.get('nombre', ''),
                'responsable_ingreso': resp_det.get('fecha_ingreso', ''),
                'responsable_cargo': resp_det.get('cargo', ''),
                'monto_contado': f.get('Monto Contado', 0),
                'monto_recibir': f.get('Monto A Recibir', 0),
                'diferencia': f.get('Diferencia Contado Vs. Recibido', 0),
                'secuencia': f.get('Secuencia De Caja'),
                'num_depositos': f.get('Número De Depósitos'),
                'estado': f.get('Estado', ''),
                'cuadre': f.get('Estado De Cuadre', ''),
                'observacion': f.get('Observación', ''),
                'evidencias': evidencias,
                'evidencias_deposito': evidencias_deposito,
                'fecha_creacion': f.get('Fecha Creación'),
                'responsable_email': (f.get('Correo (from Responsable De Caja)', [None]) or [None])[0],
            })

        response_data = {'depositos': resultado, 'total': len(resultado)}
        # Guardar en cache
        import time as _t2
        _depositos_cache[cache_key] = {'data': response_data, 'ts': _t2.time()}
        return jsonify(response_data)
    except Exception as e:
        print(f'Error en depositos_listar: {e}')
        return jsonify({'error': str(e)[:200]}), 500


@app.route('/api/depositos/resumen', methods=['GET'])
def depositos_resumen():
    """Resumen/KPIs de depositos. Cache 5 min."""
    import requests as req
    import time as _t
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    cache_key = f'resumen|{fecha_desde}|{fecha_hasta}'
    cached = _depositos_cache.get(cache_key)
    if cached and (_t.time() - cached['ts']) < _DEPOSITOS_CACHE_TTL:
        return jsonify(cached['data'])

    try:
        filtros = []
        if fecha_desde:
            filtros.append(f"IS_ON_OR_AFTER({{Fecha}}, '{fecha_desde}')")
        if fecha_hasta:
            filtros.append(f"IS_ON_OR_BEFORE({{Fecha}}, '{fecha_hasta}')")

        params = {
            'pageSize': 100,
            'fields[]': ['Fecha', 'Local', 'Monto Contado', 'Monto A Recibir',
                         'Diferencia Contado Vs. Recibido', 'Estado', 'Estado De Cuadre'],
        }
        if filtros:
            params['filterByFormula'] = 'AND(' + ','.join(filtros) + ')'

        all_records = []
        offset = None
        while True:
            if offset:
                params['offset'] = offset
            r = req.get(f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_DEPOSITOS_TABLE}',
                headers=_at_headers(), params=params, timeout=20)
            if r.status_code != 200:
                break
            data = r.json()
            all_records.extend(data.get('records', []))
            offset = data.get('offset')
            if not offset:
                break

        total_depositado = 0
        total_recibido = 0
        total_diferencia = 0
        descuadres = 0
        cuadran = 0
        por_local = {}
        pendientes = 0

        for rec in all_records:
            f = rec['fields']
            monto = f.get('Monto Contado', 0) or 0
            recibido = f.get('Monto A Recibir', 0) or 0
            dif = f.get('Diferencia Contado Vs. Recibido', 0) or 0
            total_depositado += monto
            total_recibido += recibido
            total_diferencia += abs(dif)

            if f.get('Estado De Cuadre') == 'Descuadra':
                descuadres += 1
            elif f.get('Estado De Cuadre') == 'Cuadra':
                cuadran += 1

            if f.get('Estado') not in ('Aprobado por Contabilidad',):
                pendientes += 1

            local = _resolver_local(f.get('Local', []))
            if local not in por_local:
                por_local[local] = {'monto': 0, 'depositos': 0, 'descuadres': 0}
            por_local[local]['monto'] += monto
            por_local[local]['depositos'] += 1
            if f.get('Estado De Cuadre') == 'Descuadra':
                por_local[local]['descuadres'] += 1

        response_data = {
            'total_depositos': len(all_records),
            'total_depositado': round(total_depositado, 2),
            'total_recibido': round(total_recibido, 2),
            'total_diferencia': round(total_diferencia, 2),
            'cuadran': cuadran,
            'descuadres': descuadres,
            'pendientes': pendientes,
            'por_local': por_local,
        }
        _depositos_cache[cache_key] = {'data': response_data, 'ts': _t.time()}
        return jsonify(response_data)
    except Exception as e:
        print(f'Error en depositos_resumen: {e}')
        return jsonify({'error': str(e)[:200]}), 500


@app.route('/api/depositos/trabajadores', methods=['GET'])
def depositos_trabajadores():
    """Lista trabajadores para el selector de testigo. Cache largo."""
    responsables = _cargar_responsables()
    lista = [{'id': k, 'nombre': v.get('nombre', '')} for k, v in responsables.items() if v.get('nombre')]
    lista.sort(key=lambda x: x['nombre'])
    return jsonify(lista)


@app.route('/api/depositos/verificar-lider', methods=['POST'])
def depositos_verificar_lider():
    """Lider verifica: pone monto a recibir, calcula cuadre, marca retirado."""
    _invalidar_cache_depositos()
    import requests as req
    data = request.json or {}
    record_id = data.get('id')
    monto_recibir = data.get('monto_recibir')
    observacion = data.get('observacion', 'No existe Observación')
    testigo_id = data.get('testigo', '')

    if not record_id or monto_recibir is None:
        return jsonify({'error': 'id y monto_recibir requeridos'}), 400

    try:
        fields = {
            'Estado': 'Valor retirado por líder',
            'Monto A Recibir': float(monto_recibir),
            'Fecha De Revisión líder': datetime.now().isoformat(),
            'Observación': observacion,
        }
        if testigo_id:
            fields['Testigo de conteo'] = [testigo_id]
        # Calcular cuadre
        r_get = req.get(
            f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_DEPOSITOS_TABLE}/{record_id}',
            headers=_at_headers(), params={'fields[]': ['Monto Contado']}, timeout=15)
        if r_get.status_code == 200:
            monto_contado = r_get.json().get('fields', {}).get('Monto Contado', 0) or 0
            diferencia = float(monto_contado) - float(monto_recibir)
            fields['Estado De Cuadre'] = 'Cuadra' if abs(diferencia) < 0.01 else 'Descuadra'

        r = req.patch(
            f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_DEPOSITOS_TABLE}/{record_id}',
            headers={**_at_headers(), 'Content-Type': 'application/json'},
            json={'fields': fields}, timeout=15)
        if r.status_code == 200:
            return jsonify({'ok': True, 'cuadre': fields.get('Estado De Cuadre', '')})
        return jsonify({'error': f'AirTable: {r.status_code} {r.text[:100]}'}), 500
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500


@app.route('/api/depositos/enviar-contabilidad', methods=['POST'])
def depositos_enviar_contabilidad():
    """Marca como enviado a contabilidad."""
    _invalidar_cache_depositos()
    import requests as req
    data = request.json or {}
    record_id = data.get('id')
    if not record_id:
        return jsonify({'error': 'id requerido'}), 400
    try:
        r = req.patch(
            f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_DEPOSITOS_TABLE}/{record_id}',
            headers={**_at_headers(), 'Content-Type': 'application/json'},
            json={'fields': {
                'Estado': 'Enviado a Contabilidad',
                'Fecha Envío A Contabilidad': datetime.now().isoformat(),
            }}, timeout=15)
        if r.status_code == 200:
            return jsonify({'ok': True})
        return jsonify({'error': f'AirTable: {r.status_code}'}), 500
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500


@app.route('/api/depositos/subir-respaldo', methods=['POST'])
def depositos_subir_respaldo():
    """Sube foto de respaldo: guarda en servidor + envia URL a AirTable."""
    _invalidar_cache_depositos()
    import requests as req

    record_id = request.form.get('id')
    if not record_id:
        return jsonify({'error': 'id requerido'}), 400

    archivo = request.files.get('archivo')
    if not archivo:
        return jsonify({'error': 'archivo requerido'}), 400

    try:
        contenido = archivo.read()
        filename = archivo.filename

        # 1. Guardar archivo en el servidor
        import os
        temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
        os.makedirs(temp_dir, exist_ok=True)
        safe_name = filename.replace(' ', '_').replace('/', '_')
        temp_path = os.path.join(temp_dir, f'{record_id}_{safe_name}')
        with open(temp_path, 'wb') as f:
            f.write(contenido)

        # 2. Construir URL publica (en produccion usa Render URL)
        app_url = os.environ.get('APP_URL', request.host_url.rstrip('/'))
        img_url = f'{app_url}/static/uploads/{record_id}_{safe_name}'

        # 3. Obtener attachments existentes
        r_get = req.get(
            f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_DEPOSITOS_TABLE}/{record_id}',
            headers=_at_headers(),
            params={'fields[]': ['Evidencia Del Déposito']},
            timeout=15,
        )
        existentes = []
        if r_get.status_code == 200:
            for a in r_get.json().get('fields', {}).get('Evidencia Del Déposito', []):
                if isinstance(a, dict) and a.get('url'):
                    existentes.append({'url': a['url']})

        existentes.append({'url': img_url})

        # 4. Actualizar AirTable
        r2 = req.patch(
            f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_DEPOSITOS_TABLE}/{record_id}',
            headers={**_at_headers(), 'Content-Type': 'application/json'},
            json={'fields': {
                'Evidencia Del Déposito': existentes,
                'Estado': 'Enviado a Contabilidad',
                'Fecha Envío A Contabilidad': datetime.now().isoformat(),
            }},
            timeout=15,
        )
        print(f'AirTable patch: {r2.status_code} {r2.text[:200]}')
        if r2.status_code == 200:
            return jsonify({'ok': True})
        return jsonify({'error': f'AirTable: {r2.status_code} {r2.text[:200]}'}), 500
    except Exception as e:
        print(f'Error subiendo respaldo: {e}')
        return jsonify({'error': str(e)[:200]}), 500


@app.route('/api/depositos/aprobar', methods=['POST'])
def depositos_aprobar():
    """Aprueba o rechaza un deposito (contabilidad)."""
    _invalidar_cache_depositos()
    import requests as req
    data = request.json or {}
    record_id = data.get('id')
    rechazar = data.get('rechazar', False)
    if not record_id:
        return jsonify({'error': 'id requerido'}), 400
    try:
        if rechazar:
            fields = {
                'Estado': 'Rechazado por Contabilidad',
                'Fecha Aprobado Por Contabilidad': datetime.now().isoformat(),
            }
        else:
            fields = {
                'Estado': 'Aprobado por Contabilidad',
                'Fecha Aprobado Por Contabilidad': datetime.now().isoformat(),
            }
        r = req.patch(
            f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_DEPOSITOS_TABLE}/{record_id}',
            headers={**_at_headers(), 'Content-Type': 'application/json'},
            json={'fields': fields}, timeout=15)
        if r.status_code == 200:
            return jsonify({'ok': True})
        return jsonify({'error': f'AirTable: {r.status_code}'}), 500
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500


# ============================================================
# CARGA INVENTARIO MANUAL (admin -> worker)
# ============================================================

BODEGAS_CARGA_MAP = {
    'real_audiencia': {'contifico': 'BODEGA CHIOS REAL', 'modo': 'CHIOS'},
    'floreana': {'contifico': 'BODEGA CHIOS FLOREANA', 'modo': 'CHIOS'},
    'portugal': {'contifico': 'BODEGA CHIOS PORTUGAL', 'modo': 'CHIOS'},
    'santo_cachon_real': {'contifico': 'BODEGA SANTO CACHON REAL', 'modo': 'CACHON'},
    'santo_cachon_portugal': {'contifico': 'BODEGA SANTO CACHON PORTUGAL', 'modo': 'CACHON'},
    'simon_bolon': {'contifico': 'BODEGA SIMON BOLON', 'modo': 'SIMON_BOLON'},
}

@app.route('/api/admin/cargar-inventario', methods=['POST'])
def admin_cargar_inventario():
    """Admin solicita carga de inventario para una bodega+fecha."""
    data = request.json or {}
    bodega = data.get('bodega')
    fecha = data.get('fecha')
    usuario = data.get('usuario', 'admin')

    if bodega not in BODEGAS_CARGA_MAP:
        return jsonify({'error': 'bodega invalida'}), 400
    if not fecha:
        return jsonify({'error': 'fecha requerida'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            SELECT id, estado FROM goti.carga_inventario_tareas
            WHERE bodega = %s AND fecha = %s
        """, (bodega, fecha))
        existente = cur.fetchone()

        if existente:
            if existente['estado'] in ('pendiente', 'en_proceso'):
                return jsonify({'id': existente['id'], 'estado': existente['estado'], 'reused': True})
            # Resetear si completado o error
            cur.execute("""
                UPDATE goti.carga_inventario_tareas
                SET estado='pendiente', solicitado_por=%s, solicitado_at=NOW(),
                    worker_lock=NULL, error_msg=NULL, timestamp_inicio=NULL,
                    timestamp_fin=NULL, total_productos=NULL
                WHERE id = %s
            """, (usuario, existente['id']))
            conn.commit()
            return jsonify({'id': existente['id'], 'estado': 'pendiente', 'reset': True})

        modo = BODEGAS_CARGA_MAP[bodega]['modo']
        cur.execute("""
            INSERT INTO goti.carga_inventario_tareas (bodega, fecha, modo, solicitado_por)
            VALUES (%s, %s, %s, %s) RETURNING id
        """, (bodega, fecha, modo, usuario))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'id': new_id, 'estado': 'pendiente'})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/admin/cargar-inventario/pendientes', methods=['GET'])
def admin_cargar_inventario_pendientes():
    """Worker toma tareas de carga."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401
    worker_id = request.args.get('worker_id', 'pc-finanzas')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.carga_inventario_tareas
            SET estado = 'en_proceso', worker_lock = %s, timestamp_inicio = NOW()
            WHERE id IN (
                SELECT id FROM goti.carga_inventario_tareas
                WHERE estado = 'pendiente' ORDER BY solicitado_at ASC LIMIT 1
                FOR UPDATE SKIP LOCKED
            )
            RETURNING id, bodega, fecha, modo
        """, (worker_id,))
        rows = cur.fetchall()
        conn.commit()
        return jsonify([{
            'id': r['id'], 'bodega': r['bodega'],
            'fecha': r['fecha'].isoformat() if r['fecha'] else None,
            'modo': r['modo'], 'tipo': 'carga_inventario',
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/admin/cargar-inventario/resultado', methods=['POST'])
def admin_cargar_inventario_resultado():
    """Worker reporta resultado."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401
    data = request.json or {}
    ejec_id = data.get('id')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.carga_inventario_tareas
            SET estado = %s, timestamp_fin = NOW(), total_productos = %s, error_msg = %s
            WHERE id = %s
        """, (data.get('estado', 'completado'), data.get('total_productos'), data.get('error_msg'), ejec_id))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/admin/cargar-inventario/estado/<int:ejec_id>', methods=['GET'])
def admin_cargar_inventario_estado(ejec_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM goti.carga_inventario_tareas WHERE id = %s", (ejec_id,))
        r = cur.fetchone()
        if not r: return jsonify({'error': 'no encontrado'}), 404
        return jsonify({
            'id': r['id'], 'bodega': r['bodega'], 'estado': r['estado'],
            'total_productos': r['total_productos'], 'error_msg': r['error_msg'],
        })
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


# ============================================================
# DIAS LIBRES GERENTES
# ============================================================

@app.route('/api/depositos/horario-gerentes', methods=['GET'])
def horario_gerentes_listar():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id, local, gerente, lunes, martes, miercoles, jueves, viernes, sabado, domingo FROM goti.horario_gerentes ORDER BY local")
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/depositos/horario-gerentes', methods=['POST'])
def horario_gerentes_guardar():
    data = request.json or {}
    local = data.get('local')
    gerente = data.get('gerente')
    dias = data.get('dias', {})
    if not local or not gerente:
        return jsonify({'error': 'local y gerente requeridos'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.horario_gerentes (local, gerente, lunes, martes, miercoles, jueves, viernes, sabado, domingo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (local) DO UPDATE SET
                gerente = EXCLUDED.gerente,
                lunes = EXCLUDED.lunes, martes = EXCLUDED.martes, miercoles = EXCLUDED.miercoles,
                jueves = EXCLUDED.jueves, viernes = EXCLUDED.viernes, sabado = EXCLUDED.sabado,
                domingo = EXCLUDED.domingo
        """, (local, gerente,
              dias.get('lunes', True), dias.get('martes', True), dias.get('miercoles', True),
              dias.get('jueves', True), dias.get('viernes', True), dias.get('sabado', True),
              dias.get('domingo', False)))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


# ==================== ADMIN USUARIOS ====================

SMTP_CONFIG = {
    'server': os.environ.get('SMTP_SERVER', 'smtp.gmail.com'),
    'port': int(os.environ.get('SMTP_PORT', '587')),
    'user': os.environ.get('SMTP_USER', 'ortiz.medranda@gmail.com'),
    'password': os.environ.get('SMTP_PASSWORD', 'curahoyg mkxzkbwz'),
}
APP_URL = os.environ.get('APP_URL', 'https://inventario-ciego-5bdr.onrender.com')


def _enviar_email_invitacion(email_destino, nombre, username, token):
    """Envia email con link para establecer contrasena."""
    link = f"{APP_URL}/establecer-clave?token={token}"
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;padding:30px;background:#f8fafc;border-radius:12px;">
        <div style="text-align:center;margin-bottom:24px;">
            <h1 style="color:#123450;font-size:22px;margin:0;">FOODIX - Inventario</h1>
        </div>
        <div style="background:#fff;padding:28px;border-radius:10px;border:1px solid #e2e8f0;">
            <h2 style="color:#123450;font-size:18px;margin:0 0 12px;">Hola {nombre},</h2>
            <p style="color:#475569;font-size:14px;line-height:1.6;">
                Se ha creado tu cuenta en el sistema de inventario.
                Tu usuario es: <strong style="color:#123450;">{username}</strong>
            </p>
            <p style="color:#475569;font-size:14px;line-height:1.6;">
                Haz clic en el boton para establecer tu contrasena:
            </p>
            <div style="text-align:center;margin:24px 0;">
                <a href="{link}" style="background:#123450;color:#fff;padding:14px 32px;border-radius:8px;text-decoration:none;font-weight:600;font-size:14px;display:inline-block;">
                    Crear mi contrasena
                </a>
            </div>
            <p style="color:#94a3b8;font-size:12px;line-height:1.5;">
                Este enlace es valido por 48 horas. Si no puedes hacer clic, copia y pega esta URL en tu navegador:<br>
                <span style="color:#64748b;word-break:break-all;">{link}</span>
            </p>
        </div>
        <p style="color:#94a3b8;font-size:11px;text-align:center;margin-top:16px;">
            FOODIX S.A.S. — Sistema de Inventario Ciego
        </p>
    </div>
    """
    msg = MIMEMultipart('alternative')
    msg['Subject'] = f'FOODIX Inventario — Configura tu acceso'
    msg['From'] = f'FOODIX Inventario <{SMTP_CONFIG["user"]}>'
    msg['To'] = email_destino
    msg.attach(MIMEText(html, 'html'))

    server = smtplib.SMTP(SMTP_CONFIG['server'], SMTP_CONFIG['port'], timeout=15)
    server.starttls()
    server.login(SMTP_CONFIG['user'], SMTP_CONFIG['password'])
    server.sendmail(SMTP_CONFIG['user'], email_destino, msg.as_string())
    server.quit()


def _require_admin(data):
    """Valida que quien llama sea admin (username + password en el body)."""
    if not data:
        return None, jsonify({'error': 'Sin datos'}), 400
    admin_user = data.get('admin_user', '')
    admin_pass = data.get('admin_pass', '')
    if not admin_user or not admin_pass:
        return None, jsonify({'error': 'Credenciales de admin requeridas'}), 401
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""SELECT id FROM goti.usuarios
                       WHERE username = %s AND password = %s AND rol = 'admin' AND activo = TRUE""",
                    (admin_user, admin_pass))
        row = cur.fetchone()
        if not row:
            return None, jsonify({'error': 'No autorizado'}), 403
        return conn, None, None
    except Exception:
        release_db(conn)
        raise


@app.route('/api/admin/personas', methods=['GET'])
def admin_listar_personas():
    """Devuelve lista de personas activas desde AirTable con nombre y correo."""
    try:
        personas = _obtener_personas_con_correo()
        return jsonify(personas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/usuarios', methods=['GET'])
def admin_listar_usuarios():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT u.id, u.username, u.nombre, u.rol, u.activo, u.created_at, u.email,
                   COALESCE(array_agg(ub.bodega ORDER BY ub.bodega) FILTER (WHERE ub.bodega IS NOT NULL), '{}') AS bodegas
            FROM goti.usuarios u
            LEFT JOIN goti.usuario_bodegas ub ON ub.usuario_id = u.id
            GROUP BY u.id, u.username, u.nombre, u.rol, u.activo, u.created_at, u.email
            ORDER BY u.id
        """)
        usuarios = cur.fetchall()
        return jsonify(usuarios)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/usuarios', methods=['POST'])
def admin_crear_usuario():
    data = request.json
    conn, err, code = _require_admin(data)
    if err:
        return err, code
    try:
        cur = conn.cursor()
        username = data.get('username', '').strip().lower()
        nombre = data.get('nombre', '').strip()
        password = data.get('password', '').strip()
        email = data.get('email', '').strip().lower()
        rol = data.get('rol', 'subgerente')
        bodegas = data.get('bodegas', [])
        enviar_invitacion = data.get('enviar_invitacion', False)

        if not username or not nombre:
            return jsonify({'error': 'username y nombre son obligatorios'}), 400
        if enviar_invitacion and not email:
            return jsonify({'error': 'Email es obligatorio para enviar invitacion'}), 400
        if not enviar_invitacion and not password:
            return jsonify({'error': 'Debes asignar contrasena o enviar invitacion por email'}), 400

        cur.execute("SELECT id FROM goti.usuarios WHERE username = %s", (username,))
        if cur.fetchone():
            return jsonify({'error': f'El usuario "{username}" ya existe'}), 409

        # Generar token si enviar invitacion
        token = secrets.token_urlsafe(32) if enviar_invitacion else None
        token_expires = (datetime.utcnow() + timedelta(hours=48)).isoformat() if token else None
        pwd = password if password else '__pendiente__'

        cur.execute("""
            INSERT INTO goti.usuarios (username, password, nombre, rol, activo, email, invite_token, invite_token_expires)
            VALUES (%s, %s, %s, %s, TRUE, %s, %s, %s) RETURNING id
        """, (username, pwd, nombre, rol, email or None, token, token_expires))
        new_id = cur.fetchone()['id']

        for bod in bodegas:
            cur.execute("""INSERT INTO goti.usuario_bodegas (usuario_id, bodega)
                           VALUES (%s, %s) ON CONFLICT DO NOTHING""", (new_id, bod))

        conn.commit()

        # Enviar email
        msg_extra = ''
        if enviar_invitacion and token:
            try:
                _enviar_email_invitacion(email, nombre, username, token)
                msg_extra = f' — Invitacion enviada a {email}'
            except Exception as mail_err:
                msg_extra = f' — ERROR enviando email: {str(mail_err)[:100]}'

        return jsonify({'success': True, 'id': new_id, 'message': f'Usuario {username} creado{msg_extra}'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db(conn)


@app.route('/api/admin/usuarios/<int:uid>', methods=['PUT'])
def admin_editar_usuario(uid):
    data = request.json
    conn, err, code = _require_admin(data)
    if err:
        return err, code
    try:
        cur = conn.cursor()
        username = data.get('username', '').strip().lower()
        nombre = data.get('nombre', '').strip()
        password = data.get('password', '').strip()
        rol = data.get('rol', 'subgerente')
        activo = data.get('activo', True)
        bodegas = data.get('bodegas', [])

        email = data.get('email', '').strip().lower() or None

        # Verificar que el nuevo username no exista en otro usuario
        if username:
            cur.execute("SELECT id FROM goti.usuarios WHERE username = %s AND id != %s", (username, uid))
            if cur.fetchone():
                return jsonify({'error': f'El usuario "{username}" ya existe'}), 409

        if password:
            cur.execute("""UPDATE goti.usuarios
                           SET username = %s, nombre = %s, password = %s, rol = %s, activo = %s, email = %s
                           WHERE id = %s""", (username, nombre, password, rol, activo, email, uid))
        else:
            cur.execute("""UPDATE goti.usuarios
                           SET username = %s, nombre = %s, rol = %s, activo = %s, email = %s
                           WHERE id = %s""", (username, nombre, rol, activo, email, uid))

        if cur.rowcount == 0:
            return jsonify({'error': 'Usuario no encontrado'}), 404

        cur.execute("DELETE FROM goti.usuario_bodegas WHERE usuario_id = %s", (uid,))
        for bod in bodegas:
            cur.execute("""INSERT INTO goti.usuario_bodegas (usuario_id, bodega)
                           VALUES (%s, %s) ON CONFLICT DO NOTHING""", (uid, bod))

        conn.commit()
        return jsonify({'success': True, 'message': 'Usuario actualizado'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db(conn)


@app.route('/api/admin/usuarios/<int:uid>', methods=['DELETE'])
def admin_eliminar_usuario(uid):
    data = request.json
    conn, err, code = _require_admin(data)
    if err:
        return err, code
    try:
        cur = conn.cursor()
        cur.execute("SELECT username FROM goti.usuarios WHERE id = %s", (uid,))
        row = cur.fetchone()
        if row and row['username'] == 'admin':
            return jsonify({'error': 'No se puede eliminar al administrador principal'}), 403

        cur.execute("DELETE FROM goti.usuarios WHERE id = %s", (uid,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Usuario eliminado'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db(conn)


@app.route('/api/admin/roles', methods=['GET'])
def admin_listar_roles():
    """Devuelve los modulos y permisos de cada rol."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""SELECT rol, modulo, puede_ver, puede_editar,
                       COALESCE(puede_eliminar, FALSE) as puede_eliminar
                       FROM goti.rol_modulos ORDER BY rol, modulo""")
        rows = cur.fetchall()
        result = {}
        for r in rows:
            result.setdefault(r['rol'], {})[r['modulo']] = {
                'ver': r['puede_ver'], 'editar': r['puede_editar'], 'eliminar': r['puede_eliminar']
            }
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/roles', methods=['PUT'])
def admin_guardar_roles():
    """Guarda permisos de un rol. Body: { admin_user, admin_pass, rol, modulos: { modulo: {ver,editar,eliminar} } }"""
    data = request.json
    conn, err, code = _require_admin(data)
    if err:
        return err, code
    try:
        cur = conn.cursor()
        rol = data.get('rol', '').strip().lower()
        modulos = data.get('modulos', {})
        if rol not in ('subgerente', 'supervisor', 'gerente', 'admin'):
            return jsonify({'error': 'Rol invalido'}), 400
        cur.execute("DELETE FROM goti.rol_modulos WHERE rol = %s", (rol,))
        for mod, perms in modulos.items():
            ver = perms.get('ver', False)
            editar = perms.get('editar', False)
            eliminar = perms.get('eliminar', False)
            if ver or editar or eliminar:
                cur.execute("""INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar)
                               VALUES (%s, %s, %s, %s, %s) ON CONFLICT DO NOTHING""",
                            (rol, mod, ver, editar, eliminar))
        conn.commit()
        return jsonify({'success': True, 'message': f'Permisos de {rol} actualizados'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db(conn)


@app.route('/api/admin/usuarios/<int:uid>/reenviar', methods=['POST'])
def admin_reenviar_invitacion(uid):
    """Genera nuevo token y reenvia email de invitacion."""
    data = request.json
    conn, err, code = _require_admin(data)
    if err:
        return err, code
    try:
        cur = conn.cursor()
        cur.execute("SELECT username, nombre, email FROM goti.usuarios WHERE id = %s", (uid,))
        user = cur.fetchone()
        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        if not user['email']:
            return jsonify({'error': 'El usuario no tiene email configurado'}), 400

        token = secrets.token_urlsafe(32)
        token_expires = (datetime.utcnow() + timedelta(hours=48)).isoformat()
        cur.execute("""UPDATE goti.usuarios
                       SET invite_token = %s, invite_token_expires = %s
                       WHERE id = %s""", (token, token_expires, uid))
        conn.commit()

        _enviar_email_invitacion(user['email'], user['nombre'], user['username'], token)
        return jsonify({'success': True, 'message': f'Invitacion reenviada a {user["email"]}'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db(conn)


PAGINA_ESTABLECER_CLAVE = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Establecer Contrasena - FOODIX</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Inter', sans-serif; background: #F8FAFC; min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 20px; }
        .card { background: #fff; border-radius: 16px; box-shadow: 0 4px 24px rgba(15,23,42,0.08); max-width: 420px; width: 100%; padding: 40px 32px; }
        .logo { text-align: center; margin-bottom: 28px; }
        .logo h1 { color: #123450; font-size: 24px; }
        .logo p { color: #64748B; font-size: 13px; margin-top: 4px; }
        .info { background: #EFF6FF; border: 1px solid #BFDBFE; border-radius: 10px; padding: 14px 16px; margin-bottom: 20px; }
        .info p { color: #1E40AF; font-size: 13px; line-height: 1.5; }
        .info strong { color: #123450; }
        .form-group { margin-bottom: 16px; }
        .form-group label { display: block; font-size: 12px; font-weight: 600; color: #64748B; text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 6px; }
        .form-group input { width: 100%; padding: 12px 14px; border: 1px solid #CBD5E1; border-radius: 10px; font-size: 14px; font-family: inherit; transition: border 0.2s; }
        .form-group input:focus { outline: none; border-color: #123450; box-shadow: 0 0 0 3px rgba(18,52,80,0.1); }
        .btn { width: 100%; padding: 14px; background: #123450; color: #fff; border: none; border-radius: 10px; font-size: 14px; font-weight: 600; cursor: pointer; transition: all 0.2s; }
        .btn:hover { background: #1a4a6e; }
        .btn:disabled { opacity: 0.5; cursor: not-allowed; }
        .msg { text-align: center; padding: 12px; border-radius: 8px; margin-top: 16px; font-size: 13px; display: none; }
        .msg.ok { display: block; background: #D1FAE5; color: #065F46; }
        .msg.err { display: block; background: #FEE2E2; color: #991B1B; }
        .success-card { text-align: center; }
        .success-card .icon { font-size: 48px; margin-bottom: 16px; }
        .success-card a { display: inline-block; margin-top: 20px; padding: 12px 28px; background: #123450; color: #fff; border-radius: 10px; text-decoration: none; font-weight: 600; font-size: 14px; }
        .success-card a:hover { background: #1a4a6e; }
    </style>
</head>
<body>
    <div class="card" id="form-card">
        <div class="logo">
            <h1>FOODIX</h1>
            <p>Sistema de Inventario</p>
        </div>
        <div class="info">
            <p>Hola <strong>{{ nombre }}</strong>, establece la contrasena para tu usuario <strong>{{ username }}</strong></p>
        </div>
        <form id="set-pass-form" onsubmit="return guardar(event)">
            <div class="form-group">
                <label>Nueva contrasena</label>
                <input type="password" id="pass1" placeholder="Minimo 4 caracteres" required minlength="4">
            </div>
            <div class="form-group">
                <label>Confirmar contrasena</label>
                <input type="password" id="pass2" placeholder="Repite la contrasena" required minlength="4">
            </div>
            <button type="submit" class="btn" id="btn-guardar">Establecer contrasena</button>
        </form>
        <div id="msg" class="msg"></div>
    </div>
    <div class="card success-card" id="success-card" style="display:none;">
        <div class="icon">&#10004;</div>
        <h2 style="color:#065F46;font-size:20px;">Contrasena establecida</h2>
        <p style="color:#64748B;margin-top:8px;font-size:14px;">Ya puedes iniciar sesion con tu usuario y contrasena.</p>
        <a href="/">Ir al sistema</a>
    </div>
    <script>
    async function guardar(e) {
        e.preventDefault();
        const p1 = document.getElementById('pass1').value;
        const p2 = document.getElementById('pass2').value;
        const msg = document.getElementById('msg');
        if (p1 !== p2) { msg.className = 'msg err'; msg.textContent = 'Las contrasenas no coinciden'; return false; }
        document.getElementById('btn-guardar').disabled = true;
        try {
            const r = await fetch('/api/establecer-clave', {
                method: 'POST', headers: {'Content-Type':'application/json'},
                body: JSON.stringify({token: '{{ token }}', password: p1})
            });
            const d = await r.json();
            if (r.ok && d.success) {
                document.getElementById('form-card').style.display = 'none';
                document.getElementById('success-card').style.display = 'block';
            } else {
                msg.className = 'msg err'; msg.textContent = d.error || 'Error al guardar';
                document.getElementById('btn-guardar').disabled = false;
            }
        } catch(err) {
            msg.className = 'msg err'; msg.textContent = 'Error de conexion';
            document.getElementById('btn-guardar').disabled = false;
        }
        return false;
    }
    </script>
</body>
</html>
"""

PAGINA_TOKEN_INVALIDO = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Enlace invalido - FOODIX</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Inter', sans-serif; background: #F8FAFC; min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 20px; }
        .card { background: #fff; border-radius: 16px; box-shadow: 0 4px 24px rgba(15,23,42,0.08); max-width: 420px; width: 100%; padding: 40px 32px; text-align: center; }
        .icon { font-size: 48px; margin-bottom: 16px; }
        h2 { color: #991B1B; font-size: 20px; }
        p { color: #64748B; margin-top: 8px; font-size: 14px; }
    </style>
</head>
<body>
    <div class="card">
        <div class="icon">&#10060;</div>
        <h2>Enlace invalido o expirado</h2>
        <p>Pide al administrador que te reenvie la invitacion.</p>
    </div>
</body>
</html>
"""


@app.route('/api/establecer-clave', methods=['POST'])
def api_establecer_clave():
    """Endpoint para guardar la contrasena desde el formulario publico."""
    data = request.json or {}
    token = data.get('token', '')
    password = data.get('password', '')

    if not token or not password:
        return jsonify({'error': 'Token y contrasena requeridos'}), 400
    if len(password) < 4:
        return jsonify({'error': 'La contrasena debe tener al menos 4 caracteres'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""SELECT id, invite_token_expires FROM goti.usuarios
                       WHERE invite_token = %s AND activo = TRUE""", (token,))
        user = cur.fetchone()
        if not user:
            return jsonify({'error': 'Enlace invalido'}), 404
        if user['invite_token_expires'] and user['invite_token_expires'] < datetime.utcnow():
            return jsonify({'error': 'Enlace expirado. Pide al administrador que lo reenvie.'}), 410

        cur.execute("""UPDATE goti.usuarios
                       SET password = %s, invite_token = NULL, invite_token_expires = NULL
                       WHERE id = %s""", (password, user['id']))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port, debug=False)
